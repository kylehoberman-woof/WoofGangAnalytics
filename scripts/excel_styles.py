"""
Excel styling and formatting utilities for openpyxl workbooks.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import PAW_MAGENTA, TEDDY_BROWN, LIGHT_PINK, DARK_TEAL, WHITE
from formatting import fmt_currency, fmt_pct, fmt_int


def style_header(ws, row, max_col, fill_color=DARK_TEAL):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="Inter", size=10, bold=True, color=WHITE)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_title(ws, row, text, col=1, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Inter", size=14, bold=True, color=PAW_MAGENTA)
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)


def style_subtitle(ws, row, text, col=1, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Inter", size=10, italic=True, color="666666")
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)


def auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def zebra_rows(ws, start_row, end_row, max_col):
    light = PatternFill(start_color=LIGHT_PINK, end_color=LIGHT_PINK, fill_type="solid")
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = light
