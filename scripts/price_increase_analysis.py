#!/usr/bin/env python3
"""
Price Increase Impact Analysis for Woof Gang Port Washington, NY (#264)
Supports Kyle's $5 grooming/bath price increase effective May 1.
"""

import json
import os
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# Import classification and helpers from main pipeline
from run import (
    classify_item, STORE_NAME, DATA_DIR, OUTPUT_DIR,
    PAW_MAGENTA, TEDDY_BROWN, LIGHT_PINK, DARK_TEAL, WHITE,
    _style_header, _style_title, _style_subtitle, _auto_width, _zebra_rows,
    _fmt_currency, _fmt_pct, _fmt_int,
)

# ─── Load cached data ────────────────────────────────────────────────────────

def load_data():
    cache = DATA_DIR / "all_data.json"
    with open(cache) as f:
        raw = json.load(f)

    items = raw["order_items"]
    orders = raw["orders"]

    rows = []
    for item in items:
        name = item.get("Name", "")
        sku = item.get("Sku", "")
        cls = classify_item(name, sku)
        net = float(item.get("Price", 0)) * float(item.get("Quantity", 1)) - float(item.get("Discount", 0))
        rows.append({
            "order_id": item.get("OrderId"),
            "customer_id": item.get("CustomerId"),
            "created": item.get("CreatedOn"),
            "name": name,
            "sku": sku,
            "price": float(item.get("Price", 0)),
            "quantity": float(item.get("Quantity", 1)),
            "discount": float(item.get("Discount", 0)),
            "net_sales": net,
            "salesperson": item.get("SalesPerson", ""),
            **cls,
        })

    df = pd.DataFrame(rows)
    df["created"] = pd.to_datetime(df["created"])
    df["month"] = df["created"].dt.month
    df = df[df["groom_category"] != "exclude"]

    order_rows = [{"order_id": o["OrderId"], "customer_id": o["CustomerId"],
                   "created": o["CreatedOn"], "subtotal": float(o.get("SubTotal", 0)),
                   "total": float(o.get("Total", 0))} for o in orders]
    df_orders = pd.DataFrame(order_rows)
    df_orders["created"] = pd.to_datetime(df_orders["created"])
    df_orders["month"] = df_orders["created"].dt.month

    return df, df_orders


# ─── Analysis Functions ───────────────────────────────────────────────────────

def current_pricing_matrix(df):
    """Build current pricing by service type and size."""
    core = df[df["groom_category"] == "core"].copy()
    if core.empty:
        return pd.DataFrame()

    matrix = core.groupby(["service_type", "dog_size"]).agg(
        avg_price=("price", "mean"),
        median_price=("price", "median"),
        min_price=("price", "min"),
        max_price=("price", "max"),
        appointments=("quantity", "sum"),
        revenue=("net_sales", "sum"),
    ).reset_index()
    matrix = matrix.sort_values(["service_type", "dog_size"])
    return matrix


def revenue_impact_model(df):
    """Model the revenue impact of $5 increase on core grooms and baths."""
    core = df[df["groom_category"] == "core"].copy()
    if core.empty:
        return {}

    # Annual appointments by service type
    by_type = core.groupby("service_type").agg(
        annual_appts=("quantity", "sum"),
        annual_revenue=("net_sales", "sum"),
        avg_ticket=("price", "mean"),
    ).reset_index()

    by_type["new_avg_ticket"] = by_type["avg_ticket"] + 5
    by_type["incremental_revenue"] = by_type["annual_appts"] * 5
    by_type["pct_increase"] = 5 / by_type["avg_ticket"] * 100
    by_type["new_annual_revenue"] = by_type["annual_revenue"] + by_type["incremental_revenue"]

    total_current = by_type["annual_revenue"].sum()
    total_incremental = by_type["incremental_revenue"].sum()
    total_appts = by_type["annual_appts"].sum()

    # Monthly projection (assume even distribution, but weight by actual monthly volume)
    monthly = core.groupby("month")["quantity"].sum()
    may_dec_appts = monthly[monthly.index >= 5].sum()  # May through Dec
    may_dec_incremental = may_dec_appts * 5

    # Commission impact (50% split)
    groomer_uplift = total_incremental * 0.50
    store_uplift = total_incremental * 0.50

    return {
        "by_type": by_type,
        "total_current": total_current,
        "total_incremental": total_incremental,
        "total_new": total_current + total_incremental,
        "total_appts": total_appts,
        "pct_increase": total_incremental / total_current * 100 if total_current else 0,
        "may_dec_incremental": may_dec_incremental,
        "may_dec_appts": may_dec_appts,
        "groomer_uplift_annual": groomer_uplift,
        "store_uplift_annual": store_uplift,
        "avg_groomer_uplift_per_appt": 2.50,
    }


def customer_sensitivity(df, df_orders):
    """Analyze customer segments for churn risk from price increase."""
    # Get groom customers
    groom_orders = set(df[df["is_groom"]]["order_id"].unique())
    groom_customers = set(df[df["is_groom"]]["customer_id"].unique())

    # Customer segments by visit frequency
    cust_orders = df_orders[df_orders["customer_id"].isin(groom_customers)]
    cust_visits = cust_orders.groupby("customer_id").agg(
        visits=("created", lambda x: x.dt.date.nunique()),
        total_spend=("subtotal", "sum"),
        last_visit=("created", "max"),
    ).reset_index()

    # Segment
    def segment(row):
        v = row["visits"]
        if v == 1: return "1-visit (trial)"
        elif v == 2: return "2 visits"
        elif v <= 4: return "3-4 visits"
        elif v <= 8: return "5-8 visits (loyal)"
        elif v <= 12: return "9-12 visits"
        return "13+ visits (champion)"

    def churn_risk(row):
        """Estimate churn risk from a $5 increase."""
        v = row["visits"]
        spv = row["total_spend"] / v if v > 0 else 0
        if v >= 5 and spv >= 80:
            return "Very Low"  # Loyal, high-spend — not leaving over $5
        elif v >= 3 and spv >= 60:
            return "Low"
        elif v >= 2:
            return "Moderate"
        else:
            return "Higher"  # 1-visit customers, but many never return anyway

    cust_visits["segment"] = cust_visits.apply(segment, axis=1)
    cust_visits["churn_risk"] = cust_visits.apply(churn_risk, axis=1)

    # Aggregate by segment
    seg_summary = cust_visits.groupby("segment").agg(
        customers=("customer_id", "count"),
        total_spend=("total_spend", "sum"),
        avg_spend=("total_spend", "mean"),
    ).reset_index()
    total_custs = seg_summary["customers"].sum()
    total_spend = seg_summary["total_spend"].sum()
    seg_summary["pct_customers"] = seg_summary["customers"] / total_custs * 100
    seg_summary["pct_revenue"] = seg_summary["total_spend"] / total_spend * 100

    # Churn risk aggregate
    risk_summary = cust_visits.groupby("churn_risk").agg(
        customers=("customer_id", "count"),
        total_spend=("total_spend", "sum"),
    ).reset_index()
    risk_summary["pct_customers"] = risk_summary["customers"] / total_custs * 100
    risk_summary["pct_revenue"] = risk_summary["total_spend"] / total_spend * 100

    # Key insight: What % of revenue comes from low-risk customers?
    low_risk = cust_visits[cust_visits["churn_risk"].isin(["Very Low", "Low"])]
    low_risk_revenue = low_risk["total_spend"].sum()
    low_risk_pct = low_risk_revenue / total_spend * 100 if total_spend else 0

    return {
        "segment_summary": seg_summary,
        "risk_summary": risk_summary,
        "total_groom_customers": total_custs,
        "low_risk_revenue_pct": low_risk_pct,
        "low_risk_customer_count": len(low_risk),
    }


def ticket_trend_analysis(df):
    """Show how avg ticket has trended monthly — natural price creep."""
    core = df[df["groom_category"] == "core"].copy()
    monthly = core.groupby("month").agg(
        avg_ticket=("price", "mean"),
        median_ticket=("price", "median"),
        appointments=("quantity", "sum"),
        revenue=("net_sales", "sum"),
    ).reset_index()
    monthly["month_name"] = monthly["month"].apply(lambda m: datetime(2025, m, 1).strftime("%b"))

    # Calculate H1 vs H2 comparison
    h1 = core[core["month"] <= 6]
    h2 = core[core["month"] > 6]
    h1_avg = h1["price"].mean() if not h1.empty else 0
    h2_avg = h2["price"].mean() if not h2.empty else 0

    return {
        "monthly": monthly,
        "h1_avg_ticket": h1_avg,
        "h2_avg_ticket": h2_avg,
        "trend_pct": (h2_avg - h1_avg) / h1_avg * 100 if h1_avg else 0,
    }


def competitive_benchmark(df):
    """Compare Port Washington pricing to Northville benchmark."""
    # Northville data from the reference workbook
    northville = {
        "Full Groom": {"0-20 lbs": 69.77, "21-40 lbs": 79.95, "41-75 lbs": 95.00,
                        "76-100 lbs": 109.82, "Over 100 lbs": 125.00},
        "Mini Groom": {"0-20 lbs": 54.78, "21-40 lbs": 64.90, "41-75 lbs": 77.53,
                        "76-100 lbs": 90.00, "Over 100 lbs": 100.45},
        "Luxury Bath": {"0-20 lbs": 39.68, "21-40 lbs": 50.00, "41-75 lbs": 59.62,
                         "76-100 lbs": 69.73, "Over 100 lbs": 77.13},
    }

    core = df[(df["groom_category"] == "core") & (df["dog_size"].notna())].copy()
    pw_prices = core.groupby(["service_type", "dog_size"])["price"].mean().reset_index()

    comparison = []
    for _, row in pw_prices.iterrows():
        st = row["service_type"]
        size = row["dog_size"]
        pw_price = row["price"]
        nv_price = northville.get(st, {}).get(size)
        if nv_price:
            comparison.append({
                "service_type": st,
                "dog_size": size,
                "pw_current": pw_price,
                "pw_after_increase": pw_price + 5,
                "northville": nv_price,
                "pw_vs_nv_current": pw_price - nv_price,
                "pw_vs_nv_after": (pw_price + 5) - nv_price,
            })

    return pd.DataFrame(comparison)


def groomer_impact(df):
    """Show per-groomer impact of $5 increase."""
    core = df[df["groom_category"] == "core"].copy()
    by_groomer = core.groupby("salesperson").agg(
        core_appts=("quantity", "sum"),
        core_revenue=("net_sales", "sum"),
        avg_ticket=("price", "mean"),
    ).reset_index()
    by_groomer = by_groomer[by_groomer["salesperson"] != ""]
    by_groomer["incremental_revenue"] = by_groomer["core_appts"] * 5
    by_groomer["commission_uplift"] = by_groomer["incremental_revenue"] * 0.50
    by_groomer["monthly_uplift"] = by_groomer["commission_uplift"] / 12
    by_groomer = by_groomer.sort_values("core_revenue", ascending=False)
    return by_groomer


def upsell_opportunity(df):
    """Analyze SPA/add-on attach rates to find revenue beyond price increase."""
    core = df[df["groom_category"] == "core"]
    spa = df[df["groom_category"] == "spa"]
    addon = df[df["groom_category"] == "addon"]

    core_orders = set(core["order_id"].unique())
    spa_orders = set(spa["order_id"].unique())
    addon_orders = set(addon["order_id"].unique())

    core_count = len(core_orders)
    spa_count = len(core_orders & spa_orders)
    addon_count = len(core_orders & addon_orders)
    any_upsell = len(core_orders & (spa_orders | addon_orders))

    spa_rate = spa_count / core_count * 100 if core_count else 0
    addon_rate = addon_count / core_count * 100 if core_count else 0
    upsell_rate = any_upsell / core_count * 100 if core_count else 0

    # Revenue from upsells
    spa_rev = spa["net_sales"].sum()
    addon_rev = addon["net_sales"].sum()
    avg_spa = spa_rev / spa_count if spa_count else 0
    avg_addon = addon_rev / addon_count if addon_count else 0

    # If we improve attach rate by 5 pts, how much more revenue?
    target_upsell = upsell_rate + 5
    additional_upsell_orders = core_count * 0.05
    avg_upsell_value = (spa_rev + addon_rev) / any_upsell if any_upsell else 0
    upsell_opportunity = additional_upsell_orders * avg_upsell_value

    return {
        "core_appointments": core_count,
        "spa_attach_rate": spa_rate,
        "addon_attach_rate": addon_rate,
        "overall_upsell_rate": upsell_rate,
        "spa_revenue": spa_rev,
        "addon_revenue": addon_rev,
        "avg_spa_value": avg_spa,
        "avg_addon_value": avg_addon,
        "target_upsell_rate": target_upsell,
        "upsell_revenue_opportunity": upsell_opportunity,
    }


# ─── Excel Generator ─────────────────────────────────────────────────────────

def generate_report(df, df_orders):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Executive Briefing ──
    ws = wb.create_sheet("Executive Briefing")
    _style_title(ws, 1, "Price Increase Impact Analysis", merge_to=6)
    _style_subtitle(ws, 2, f"{STORE_NAME} — $5 Groom/Bath Increase Effective May 1", merge_to=6)

    r = 4
    ws.cell(row=r, column=1, value="Prepared for Kyle — March 2026").font = Font(name="Inter", size=10, italic=True, color="888888")

    r = 6
    impact = revenue_impact_model(df)
    ws.cell(row=r, column=1, value="Key Financial Impacts").font = Font(name="Inter", size=13, bold=True, color=TEDDY_BROWN)

    metrics = [
        ("Current Annual Core Groom Revenue", _fmt_currency(impact["total_current"])),
        ("Annual Incremental Revenue from $5 Increase", _fmt_currency(impact["total_incremental"])),
        ("New Annual Core Groom Revenue", _fmt_currency(impact["total_new"])),
        ("Revenue Lift", _fmt_pct(impact["pct_increase"])),
        ("", ""),
        ("May–Dec 2025 Projected Incremental (8 months)", _fmt_currency(impact["may_dec_incremental"])),
        ("Core Groom Appointments (Annual)", _fmt_int(impact["total_appts"])),
        ("", ""),
        ("Groomer Commission Uplift (Annual, at 50%)", _fmt_currency(impact["groomer_uplift_annual"])),
        ("Store Net Uplift (Annual, at 50%)", _fmt_currency(impact["store_uplift_annual"])),
        ("Per-Appointment Groomer Boost", "$2.50"),
    ]

    r += 1
    for i, h in enumerate(["Metric", "Value"], 1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, 2)

    for metric, val in metrics:
        r += 1
        ws.cell(row=r, column=1, value=metric).font = Font(name="Inter", size=10,
            bold=(metric != "" and "Current" not in metric and "Appointments" not in metric))
        c = ws.cell(row=r, column=2, value=val)
        c.font = Font(name="Inter", size=10, bold=True)
        if "Incremental" in metric or "Lift" in metric or "Uplift" in metric or "Boost" in metric:
            c.font = Font(name="Inter", size=10, bold=True, color="2E7D32")

    # Customer risk summary
    r += 3
    sens = customer_sensitivity(df, df_orders)
    ws.cell(row=r, column=1, value="Customer Retention Risk").font = Font(name="Inter", size=13, bold=True, color=TEDDY_BROWN)
    r += 1
    ws.cell(row=r, column=1, value=f"{_fmt_pct(sens['low_risk_revenue_pct'])} of grooming revenue comes from low-churn-risk customers (3+ visits, $60+ avg spend).").font = Font(name="Inter", size=11)
    r += 1
    ws.cell(row=r, column=1, value=f"A $5 increase on a $90+ avg ticket is a {5/90*100:.1f}% change — well within normal year-over-year adjustments.").font = Font(name="Inter", size=10, italic=True, color="666666")

    r += 2
    for i, h in enumerate(["Churn Risk Level", "Customers", "% of Customers", "Revenue", "% of Revenue"], 1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, 5)

    risk_order = ["Very Low", "Low", "Moderate", "Higher"]
    risk_colors = {"Very Low": "C8E6C9", "Low": "DCEDC8", "Moderate": "FFF9C4", "Higher": "FFCDD2"}
    risk_df = sens["risk_summary"]
    risk_df["_order"] = risk_df["churn_risk"].map({s: i for i, s in enumerate(risk_order)})
    risk_df = risk_df.sort_values("_order")

    for _, row in risk_df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=row["churn_risk"])
        ws.cell(row=r, column=2, value=_fmt_int(row["customers"]))
        ws.cell(row=r, column=3, value=_fmt_pct(row["pct_customers"]))
        ws.cell(row=r, column=4, value=_fmt_currency(row["total_spend"]))
        ws.cell(row=r, column=5, value=_fmt_pct(row["pct_revenue"]))
        color = risk_colors.get(row["churn_risk"], "FFFFFF")
        ws.cell(row=r, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Upsell companion opportunity
    r += 3
    upsell = upsell_opportunity(df)
    ws.cell(row=r, column=1, value="Upsell Opportunity (Companion Strategy)").font = Font(name="Inter", size=13, bold=True, color=TEDDY_BROWN)
    r += 1
    ws.cell(row=r, column=1, value="Pair the price increase with stronger SPA/add-on upselling to offset perception and grow total ticket.").font = Font(name="Inter", size=10, italic=True, color="666666")
    r += 1

    upsell_metrics = [
        ("Current SPA Attach Rate", _fmt_pct(upsell["spa_attach_rate"])),
        ("Current Add-On Attach Rate", _fmt_pct(upsell["addon_attach_rate"])),
        ("Overall Upsell Rate", _fmt_pct(upsell["overall_upsell_rate"])),
        ("Avg SPA Upgrade Value", _fmt_currency(upsell["avg_spa_value"])),
        ("Avg Add-On Value", _fmt_currency(upsell["avg_addon_value"])),
        ("", ""),
        ("If upsell rate improves 5 pts → additional revenue", _fmt_currency(upsell["upsell_revenue_opportunity"])),
        ("Combined with $5 increase → total annual uplift",
         _fmt_currency(impact["total_incremental"] + upsell["upsell_revenue_opportunity"])),
    ]
    for metric, val in upsell_metrics:
        r += 1
        ws.cell(row=r, column=1, value=metric).font = Font(name="Inter", size=10)
        ws.cell(row=r, column=2, value=val).font = Font(name="Inter", size=10, bold=True)

    _auto_width(ws)

    # ── Sheet 2: Pricing Matrix ──
    ws2 = wb.create_sheet("Current Pricing Matrix")
    _style_title(ws2, 1, "Current Pricing by Service & Size", merge_to=8)
    _style_subtitle(ws2, 2, "Based on Jan–Dec 2025 transaction data", merge_to=8)

    matrix = current_pricing_matrix(df)
    r = 4
    headers = ["Service Type", "Dog Size", "Avg Price", "After +$5", "Median Price", "Min", "Max", "Appointments", "Revenue"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=r, column=i, value=h)
    _style_header(ws2, r, len(headers))

    for _, row in matrix.iterrows():
        r += 1
        ws2.cell(row=r, column=1, value=row["service_type"])
        ws2.cell(row=r, column=2, value=row["dog_size"] if pd.notna(row["dog_size"]) else "Mixed/Flat")
        ws2.cell(row=r, column=3, value=_fmt_currency(row["avg_price"]))
        ws2.cell(row=r, column=4, value=_fmt_currency(row["avg_price"] + 5))
        ws2.cell(row=r, column=4).font = Font(name="Inter", size=10, bold=True, color="2E7D32")
        ws2.cell(row=r, column=5, value=_fmt_currency(row["median_price"]))
        ws2.cell(row=r, column=6, value=_fmt_currency(row["min_price"]))
        ws2.cell(row=r, column=7, value=_fmt_currency(row["max_price"]))
        ws2.cell(row=r, column=8, value=_fmt_int(row["appointments"]))
        ws2.cell(row=r, column=9, value=_fmt_currency(row["revenue"]))
    _zebra_rows(ws2, 5, r, len(headers))
    _auto_width(ws2)

    # ── Sheet 3: Revenue Impact by Service ──
    ws3 = wb.create_sheet("Revenue Impact Model")
    _style_title(ws3, 1, "Revenue Impact: $5 Increase by Service Type", merge_to=8)
    _style_subtitle(ws3, 2, "Projected annual impact assuming volume holds steady", merge_to=8)

    by_type = impact["by_type"]
    r = 4
    headers = ["Service Type", "Annual Appts", "Current Avg", "New Avg (+$5)", "% Increase",
               "Current Revenue", "Incremental", "New Revenue"]
    for i, h in enumerate(headers, 1):
        ws3.cell(row=r, column=i, value=h)
    _style_header(ws3, r, len(headers))

    for _, row in by_type.sort_values("annual_revenue", ascending=False).iterrows():
        r += 1
        ws3.cell(row=r, column=1, value=row["service_type"])
        ws3.cell(row=r, column=2, value=_fmt_int(row["annual_appts"]))
        ws3.cell(row=r, column=3, value=_fmt_currency(row["avg_ticket"]))
        ws3.cell(row=r, column=4, value=_fmt_currency(row["new_avg_ticket"]))
        ws3.cell(row=r, column=5, value=_fmt_pct(row["pct_increase"]))
        ws3.cell(row=r, column=6, value=_fmt_currency(row["annual_revenue"]))
        ws3.cell(row=r, column=7, value=_fmt_currency(row["incremental_revenue"]))
        ws3.cell(row=r, column=7).font = Font(name="Inter", size=10, bold=True, color="2E7D32")
        ws3.cell(row=r, column=8, value=_fmt_currency(row["new_annual_revenue"]))

    # Totals
    r += 1
    ws3.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(row=r, column=2, value=_fmt_int(by_type["annual_appts"].sum())).font = Font(bold=True)
    ws3.cell(row=r, column=6, value=_fmt_currency(by_type["annual_revenue"].sum())).font = Font(bold=True)
    ws3.cell(row=r, column=7, value=_fmt_currency(by_type["incremental_revenue"].sum())).font = Font(bold=True, color="2E7D32")
    ws3.cell(row=r, column=8, value=_fmt_currency(by_type["new_annual_revenue"].sum())).font = Font(bold=True)

    _zebra_rows(ws3, 5, r - 1, len(headers))
    _auto_width(ws3)

    # ── Sheet 4: Competitive Benchmark ──
    ws4 = wb.create_sheet("Competitive Benchmark")
    _style_title(ws4, 1, "Port Washington vs. Northville, MI (#401) Pricing", merge_to=7)
    _style_subtitle(ws4, 2, "Comparison shows where PW sits relative to another Woof Gang location", merge_to=7)

    benchmark = competitive_benchmark(df)
    r = 4
    if not benchmark.empty:
        headers = ["Service Type", "Dog Size", "PW Current", "PW After +$5", "Northville", "PW vs NV (Current)", "PW vs NV (After)"]
        for i, h in enumerate(headers, 1):
            ws4.cell(row=r, column=i, value=h)
        _style_header(ws4, r, len(headers))

        for _, row in benchmark.iterrows():
            r += 1
            ws4.cell(row=r, column=1, value=row["service_type"])
            ws4.cell(row=r, column=2, value=row["dog_size"])
            ws4.cell(row=r, column=3, value=_fmt_currency(row["pw_current"]))
            ws4.cell(row=r, column=4, value=_fmt_currency(row["pw_after_increase"]))
            ws4.cell(row=r, column=5, value=_fmt_currency(row["northville"]))
            diff = row["pw_vs_nv_current"]
            ws4.cell(row=r, column=6, value=f"{'+' if diff >= 0 else ''}{_fmt_currency(diff)}")
            ws4.cell(row=r, column=6).font = Font(color="2E7D32" if diff >= 0 else "C62828")
            diff_after = row["pw_vs_nv_after"]
            ws4.cell(row=r, column=7, value=f"{'+' if diff_after >= 0 else ''}{_fmt_currency(diff_after)}")
            ws4.cell(row=r, column=7).font = Font(color="2E7D32" if diff_after >= 0 else "C62828")

        _zebra_rows(ws4, 5, r, len(headers))

        r += 2
        ws4.cell(row=r, column=1, value="Note: Port Washington is a premium NY metro market. Higher pricing is expected and supported by market demographics.").font = Font(name="Inter", size=10, italic=True, color="666666")

    _auto_width(ws4)

    # ── Sheet 5: Groomer Commission Impact ──
    ws5 = wb.create_sheet("Groomer Impact")
    _style_title(ws5, 1, "Per-Groomer Commission Impact from $5 Increase", merge_to=7)
    _style_subtitle(ws5, 2, "Talking point for team meeting: every groomer benefits", merge_to=7)

    groomer = groomer_impact(df)
    r = 4
    headers = ["Groomer", "Annual Appts", "Current Revenue", "Current Avg", "Annual Commission Uplift", "Monthly Uplift"]
    for i, h in enumerate(headers, 1):
        ws5.cell(row=r, column=i, value=h)
    _style_header(ws5, r, len(headers))

    for _, row in groomer.iterrows():
        r += 1
        ws5.cell(row=r, column=1, value=row["salesperson"])
        ws5.cell(row=r, column=2, value=_fmt_int(row["core_appts"]))
        ws5.cell(row=r, column=3, value=_fmt_currency(row["core_revenue"]))
        ws5.cell(row=r, column=4, value=_fmt_currency(row["avg_ticket"]))
        ws5.cell(row=r, column=5, value=_fmt_currency(row["commission_uplift"]))
        ws5.cell(row=r, column=5).font = Font(name="Inter", size=10, bold=True, color="2E7D32")
        ws5.cell(row=r, column=6, value=_fmt_currency(row["monthly_uplift"]))
        ws5.cell(row=r, column=6).font = Font(name="Inter", size=10, bold=True, color="2E7D32")

    _zebra_rows(ws5, 5, r, len(headers))

    r += 1
    ws5.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws5.cell(row=r, column=2, value=_fmt_int(groomer["core_appts"].sum())).font = Font(bold=True)
    ws5.cell(row=r, column=5, value=_fmt_currency(groomer["commission_uplift"].sum())).font = Font(bold=True, color="2E7D32")
    ws5.cell(row=r, column=6, value=_fmt_currency(groomer["monthly_uplift"].sum())).font = Font(bold=True, color="2E7D32")

    r += 2
    ws5.cell(row=r, column=1, value="Key talking point: \"This increase means an extra $X/month in your pocket — we're investing in YOU.\"").font = Font(name="Inter", size=10, italic=True, color=PAW_MAGENTA)

    _auto_width(ws5)

    # ── Sheet 6: Customer Segments ──
    ws6 = wb.create_sheet("Customer Segments")
    _style_title(ws6, 1, "Grooming Customer Segments & Retention Risk", merge_to=6)
    _style_subtitle(ws6, 2, "Who are your customers and how likely are they to stay?", merge_to=6)

    r = 4
    ws6.cell(row=r, column=1, value="Visit Frequency Distribution").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    r += 1
    for i, h in enumerate(["Segment", "Customers", "% of Customers", "Revenue", "% of Revenue", "Avg Spend"], 1):
        ws6.cell(row=r, column=i, value=h)
    _style_header(ws6, r, 6)

    seg_order = ["1-visit (trial)", "2 visits", "3-4 visits", "5-8 visits (loyal)", "9-12 visits", "13+ visits (champion)"]
    seg_df = sens["segment_summary"]
    seg_df["_order"] = seg_df["segment"].map({s: i for i, s in enumerate(seg_order)})
    seg_df = seg_df.sort_values("_order")

    for _, row in seg_df.iterrows():
        r += 1
        ws6.cell(row=r, column=1, value=row["segment"])
        ws6.cell(row=r, column=2, value=_fmt_int(row["customers"]))
        ws6.cell(row=r, column=3, value=_fmt_pct(row["pct_customers"]))
        ws6.cell(row=r, column=4, value=_fmt_currency(row["total_spend"]))
        ws6.cell(row=r, column=5, value=_fmt_pct(row["pct_revenue"]))
        ws6.cell(row=r, column=6, value=_fmt_currency(row["avg_spend"]))
    _zebra_rows(ws6, 6, r, 6)

    r += 3
    ws6.cell(row=r, column=1, value="Key Insight").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    r += 1
    ws6.cell(row=r, column=1, value=f"Customers with 3+ visits represent the bulk of revenue and have the lowest churn risk from a $5 increase.").font = Font(name="Inter", size=10)
    r += 1
    ws6.cell(row=r, column=1, value=f"1-visit trial customers have a high natural attrition rate regardless of pricing — most never return even without a price change.").font = Font(name="Inter", size=10)

    _auto_width(ws6)

    # ── Sheet 7: Ticket Trend ──
    ws7 = wb.create_sheet("Ticket Trend")
    _style_title(ws7, 1, "Average Core Groom Ticket — Monthly Trend", merge_to=6)
    _style_subtitle(ws7, 2, "Shows natural ticket growth over the year (before any formal increase)", merge_to=6)

    trend = ticket_trend_analysis(df)
    r = 4
    for i, h in enumerate(["Month", "Avg Ticket", "Median Ticket", "Appointments", "Revenue"], 1):
        ws7.cell(row=r, column=i, value=h)
    _style_header(ws7, r, 5)

    for _, row in trend["monthly"].iterrows():
        r += 1
        ws7.cell(row=r, column=1, value=row["month_name"])
        ws7.cell(row=r, column=2, value=_fmt_currency(row["avg_ticket"]))
        ws7.cell(row=r, column=3, value=_fmt_currency(row["median_ticket"]))
        ws7.cell(row=r, column=4, value=_fmt_int(row["appointments"]))
        ws7.cell(row=r, column=5, value=_fmt_currency(row["revenue"]))
    _zebra_rows(ws7, 5, r, 5)

    r += 2
    ws7.cell(row=r, column=1, value=f"H1 Avg Ticket: {_fmt_currency(trend['h1_avg_ticket'])}  |  H2 Avg Ticket: {_fmt_currency(trend['h2_avg_ticket'])}  |  Natural Trend: {'+' if trend['trend_pct'] >= 0 else ''}{trend['trend_pct']:.1f}%").font = Font(name="Inter", size=10, bold=True)

    _auto_width(ws7)

    # ── Sheet 8: Rollout Talking Points ──
    ws8 = wb.create_sheet("Rollout Talking Points")
    _style_title(ws8, 1, "Data-Backed Talking Points for Team & Clients", merge_to=6)
    _style_subtitle(ws8, 2, "Use these in the Week 6 team meeting and client conversations", merge_to=6)

    r = 4
    ws8.cell(row=r, column=1, value="For the Team Meeting (Week 6)").font = Font(name="Inter", size=13, bold=True, color=TEDDY_BROWN)

    team_points = [
        f"This $5 increase translates to {_fmt_currency(impact['total_incremental'])} in additional annual revenue for the store.",
        f"At your 50% commission, that's an extra $2.50 per appointment in your pocket.",
        f"For a groomer doing 50 dogs/week, that's ~$125/week or ~$500/month more in commission.",
        f"Our average core groom ticket is already {_fmt_currency(df[df['groom_category']=='core']['price'].mean())} — this is a {5/df[df['groom_category']=='core']['price'].mean()*100:.1f}% adjustment.",
        f"We serve {sens['total_groom_customers']:,} unique grooming customers. {_fmt_pct(sens['low_risk_revenue_pct'])} of our revenue comes from loyal, low-risk customers.",
        f"This increase keeps us competitive as a premium grooming destination in the NY metro area.",
    ]

    for pt in team_points:
        r += 1
        ws8.cell(row=r, column=1, value=f"• {pt}").font = Font(name="Inter", size=10)

    r += 2
    ws8.cell(row=r, column=1, value="For Client Conversations (Weeks 4-2)").font = Font(name="Inter", size=13, bold=True, color=TEDDY_BROWN)

    client_points = [
        "\"We're adjusting our grooming prices by $5 starting May 1 — this reflects our continued investment in training, safety, and the quality of care your dog receives.\"",
        "\"Our team has grown significantly this year, and this adjustment helps us retain our best groomers and maintain the experience you expect.\"",
        "\"We wanted to give you plenty of notice — no surprises. We really value having you as part of the Woof Gang family.\"",
        "\"If you'd like to lock in current pricing, you can pre-book your next few appointments before May 1.\"",
    ]

    for pt in client_points:
        r += 1
        ws8.cell(row=r, column=1, value=f"• {pt}").font = Font(name="Inter", size=10, italic=True)

    r += 2
    ws8.cell(row=r, column=1, value="Objection Handling").font = Font(name="Inter", size=13, bold=True, color=TEDDY_BROWN)

    objections = [
        ("\"That's too expensive.\"",
         f"\"I understand. Our average groom is {_fmt_currency(df[df['groom_category']=='core']['price'].mean())} — this is less than a {5/df[df['groom_category']=='core']['price'].mean()*100:.1f}% change. We believe the quality and safety of care your dog receives here is worth it.\""),
        ("\"I can find cheaper.\"",
         "\"You absolutely can. What sets us apart is our certified groomers, premium products, and the trust you have with your groomer. We'd love to keep taking care of [dog's name].\""),
        ("\"Why now?\"",
         "\"Operating costs — especially retaining great staff — have increased. This adjustment helps us keep the team you know and love.\""),
    ]

    for obj, response in objections:
        r += 1
        ws8.cell(row=r, column=1, value=obj).font = Font(name="Inter", size=10, bold=True)
        r += 1
        ws8.cell(row=r, column=1, value=f"  → {response}").font = Font(name="Inter", size=10, italic=True, color="444444")
        r += 1

    _auto_width(ws8, min_width=12, max_width=120)

    # Save
    output = OUTPUT_DIR / "WoofGang_PortWashington_PriceIncrease_Analysis.xlsx"
    wb.save(output)
    print(f"\nREPORT SAVED: {output}")
    print(f"Sheets: {wb.sheetnames}")
    return output


# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading cached FranPOS data...")
    df, df_orders = load_data()
    print(f"Items: {len(df)}, Orders: {len(df_orders)}")
    print("\nGenerating Price Increase Impact Analysis...")
    output = generate_report(df, df_orders)
    print("Done.")
