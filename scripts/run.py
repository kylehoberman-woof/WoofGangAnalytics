#!/usr/bin/env python3
"""
Woof Gang Store Performance Analysis Pipeline
Pulls data from FranPOS API, transforms, and generates Excel workbook.
"""

import httpx
import json
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
import math
import re

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# ─── Config ──────────────────────────────────────────────────────────────────

TOKEN = "FAF11AD5771F74C3ABF5D6FB2965A21BF4F9A93B5FB3ED2AE7DF66B53C0AE23B058586C983CE6B53660530838748D490D6E10291F158DF791B1BA278FEDAA844"
BASE_URL = "https://publicapi.franpos.com"
LOCATION_ID = 203698
STORE_NAME = "Woof Gang Bakery & Grooming -- Port Washington, NY (#264)"
# ET is UTC-5 (EST) or UTC-4 (EDT). FranPOS stores in UTC.
# We fetch +1 day buffer on each side to catch timezone boundary transactions.
ET_BUFFER_DAYS = 2
START_DATE = "2024-09-26"
END_DATE = "2026-12-31"
PAGE_SIZE = 500
DATA_DIR = Path(__file__).parent.parent / "port-washington" / "data"
OUTPUT_DIR = Path(__file__).parent.parent / "port-washington"

# Brand colors
PAW_MAGENTA = "C4276E"
TEDDY_BROWN = "6B3520"
LIGHT_PINK = "FDF0F5"
DARK_TEAL = "1B6B6B"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"

# ─── API Extraction ──────────────────────────────────────────────────────────

def api_get(endpoint, params=None, timeout=60):
    """Make authenticated GET request to FranPOS API."""
    p = params or {}
    p["Token"] = TOKEN
    r = httpx.get(f"{BASE_URL}/{endpoint}", params=p, timeout=timeout)
    r.raise_for_status()
    ct = r.headers.get("content-type", "")
    if "json" in ct:
        return r.json()
    return r.text


def api_post(endpoint, body, params=None, timeout=60):
    """Make authenticated POST request."""
    p = params or {}
    p["Token"] = TOKEN
    r = httpx.post(f"{BASE_URL}/{endpoint}", params=p, json=body, timeout=timeout)
    r.raise_for_status()
    return r.json()


def extract_paginated(endpoint_template, start, end, label="data", id_field="OrderItemId"):
    """Extract paginated data one calendar month at a time to prevent boundary gaps."""
    all_data = []
    _seen = set()

    start_dt = datetime.strptime(start, "%Y-%m-%d")
    end_dt = datetime.strptime(end, "%Y-%m-%d")

    # Build list of month-start dates
    windows = []
    current = start_dt.replace(day=1)
    while current <= end_dt:
        # Window goes from the 1st to last day of the month
        if current.month == 12:
            next_month = current.replace(year=current.year + 1, month=1, day=1)
        else:
            next_month = current.replace(month=current.month + 1, day=1)
        window_end = min(next_month - timedelta(days=1), end_dt)
        # Clamp start to the requested start date
        window_start = max(current, start_dt)
        windows.append((window_start, window_end))
        current = next_month

    for window_start, window_end in windows:
        # Add buffer days on each side to catch UTC timezone boundary transactions
        fetch_start = window_start - timedelta(days=ET_BUFFER_DAYS)
        fetch_end = window_end + timedelta(days=ET_BUFFER_DAYS)
        days = (fetch_end - fetch_start).days + 1
        from_date = fetch_start.strftime("%Y-%m-%d")
        page = 1
        total_pages = 1

        while page <= total_pages:
            endpoint = endpoint_template.format(
                days=days, page=page, from_date=from_date, location_id=LOCATION_ID
            )
            print(f"  [{label}] {from_date} +{days}d, page {page}...", end="", flush=True)
            try:
                result = api_get(endpoint)
                if isinstance(result, dict):
                    items = result.get("data", [])
                    total_pages = result.get("pages", 1)
                else:
                    items = []
                if not items and page > 1:
                    print(f" 0 items — done with window")
                    break
                added = 0
                for item in items:
                    key = item.get(id_field) or item.get("OrderId") or item.get("Id")
                    if key is not None:
                        if key not in _seen:
                            _seen.add(key)
                            all_data.append(item)
                            added += 1
                    else:
                        all_data.append(item)
                        added += 1
                print(f" {len(items)} items ({added} new) (page {page}/{total_pages})")
            except Exception as e:
                print(f" ERROR: {e}")
                time.sleep(2)
                try:
                    result = api_get(endpoint)
                    if isinstance(result, dict):
                        items = result.get("data", [])
                        total_pages = result.get("pages", 1)
                    else:
                        items = []
                    added = 0
                    for item in items:
                        key = item.get(id_field) or item.get("OrderId") or item.get("Id")
                        if key is not None:
                            if key not in _seen:
                                _seen.add(key)
                                all_data.append(item)
                                added += 1
                        else:
                            all_data.append(item)
                            added += 1
                    print(f"  RETRY OK: {len(items)} items ({added} new)")
                except Exception as e2:
                    print(f"  RETRY FAILED: {e2}")
            page += 1
            time.sleep(0.3)

    print(f"  Total unique records: {len(all_data)}")
    return all_data


def extract_all_data():
    """Extract all needed data from FranPOS API, with disk caching."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    data = {}

    # Incremental fetch logic
    cache_file = DATA_DIR / "all_data.json"
    INCREMENTAL_DAYS = 60

    if cache_file.exists():
        print(f"Cache found - running incremental update (last {INCREMENTAL_DAYS} days)...")
        with open(cache_file) as f:
            cached = json.load(f)

        # If cache is fresh (< 2 hours) and has data, skip fetch entirely
        age_hours = (time.time() - cache_file.stat().st_mtime) / 3600
        if age_hours < 2 and len(cached.get("order_items", [])) > 1000:
            print(f"Cache is only {age_hours:.1f}h old - skipping fetch")
            return cached

        inc_start = (datetime.today() - timedelta(days=INCREMENTAL_DAYS)).strftime("%Y-%m-%d")

        print(f"  Incremental window: {inc_start} to today")

        print("\n[1/6] Location info...")
        loc = api_get("api/getLocationsInfo")
        cached["location"] = loc
        print(f"  Store: {loc['Data'][0]['CompanyName']}")

        print("\n[2/6] Employees...")
        emps = api_get("api/dictionary/employees")
        cached["employees"] = emps
        print(f"  {len(emps)} employees")

        print("\n[3/6] Order items (incremental)...")
        new_items = extract_paginated(
            "api/datadump/v2/orderitems/{days}/{page}/{from_date}/{location_id}",
            inc_start, END_DATE, "order_items", id_field="OrderItemId"
        )
        existing_item_ids = set()
        kept_items = []
        for item in cached.get("order_items", []):
            d = str(item.get("CreatedOn", ""))[:10]
            if d < inc_start:
                kept_items.append(item)
                existing_item_ids.add(item.get("OrderItemId"))
        for item in new_items:
            key = item.get("OrderItemId")
            if key not in existing_item_ids:
                kept_items.append(item)
                existing_item_ids.add(key)
        cached["order_items"] = kept_items

        print("\n[4/6] Orders (incremental)...")
        new_orders = extract_paginated(
            "api/datadump/v1/orders/{days}/{page}/{from_date}/{location_id}",
            inc_start, END_DATE, "orders", id_field="OrderId"
        )
        existing_order_ids = set()
        kept_orders = []
        for order in cached.get("orders", []):
            d = str(order.get("CreatedOn", ""))[:10]
            if d < inc_start:
                kept_orders.append(order)
                existing_order_ids.add(order.get("OrderId"))
        for order in new_orders:
            key = order.get("OrderId")
            if key not in existing_order_ids:
                kept_orders.append(order)
                existing_order_ids.add(key)
        cached["orders"] = kept_orders

        print(f"  After merge: {len(kept_items)} items, {len(kept_orders)} orders")

        print("\n[5/6] Time clocks (incremental)...")
        inc_start_dt = datetime.strptime(inc_start, "%Y-%m-%d")
        all_clocks = [c for c in cached.get("time_clocks", [])
                      if str(c.get("TimeIn",""))[:10] < inc_start]
        current = inc_start_dt
        while current <= datetime.today():
            window_end = min(current + timedelta(days=29), datetime.today())
            sd = current.strftime("%Y-%m-%dT00:00:00")
            ed = window_end.strftime("%Y-%m-%dT23:59:59")
            try:
                tc = api_post("api/report/timeClocks", {"startDate": sd, "endDate": ed, "pageIndex": 0, "pageSize": 200})
                records = tc if isinstance(tc, list) else tc.get("data", [])
                all_clocks.extend(records)
                print(f"  Clocks {current.strftime('%b')}: {len(records)} records")
            except Exception as e:
                print(f"  Clocks {current.strftime('%b')}: ERROR {e}")
            current = window_end + timedelta(days=1)
        cached["time_clocks"] = all_clocks

        print("\n[6/6] Categories...")
        cats = api_get("api/getCategoriesByCompany/1")
        cached["categories"] = cats
        print(f"  {len(cats)} categories")

        with open(cache_file, "w") as f:
            json.dump(cached, f)
        print("Incremental cache saved.")
        return cached

    print("=" * 60)
    print("EXTRACTING DATA FROM FRANPOS API")
    print("=" * 60)

    # 1. Location info
    print("\n[1/6] Location info...")
    loc = api_get("api/getLocationsInfo")
    data["location"] = loc
    print(f"  Store: {loc['Data'][0]['CompanyName']}")

    # 2. Employees
    print("\n[2/6] Employees...")
    emps = api_get("api/dictionary/employees")
    data["employees"] = emps
    print(f"  {len(emps)} employees")

    # 3. Order items (line-item level) - THE BIG ONE
    print("\n[3/6] Order items (line-item detail)...")
    items = extract_paginated(
        "api/datadump/v2/orderitems/{days}/{page}/{from_date}/{location_id}",
        START_DATE, END_DATE, "order_items", id_field="OrderItemId"
    )
    data["order_items"] = items
    print(f"  TOTAL: {len(items)} line items")

    # 4. Orders (order level)
    print("\n[4/6] Orders (transaction level)...")
    orders = extract_paginated(
        "api/datadump/v1/orders/{days}/{page}/{from_date}/{location_id}",
        START_DATE, END_DATE, "orders", id_field="OrderId"
    )
    data["orders"] = orders
    print(f"  TOTAL: {len(orders)} orders")


    # AUTO GAP DETECTION & PATCH
    print("\n[Gap detection] Checking for missing data...")
    from collections import defaultdict as _dd
    from datetime import date as _date, datetime as _dtt, timedelta as _td

    KNOWN_CLOSURES = {
        "2024-09-26","2024-11-28","2024-12-25","2025-01-01","2025-07-04",
        "2025-09-01","2025-11-27","2025-12-25","2026-01-01","2025-12-27",
        "2026-01-25","2026-01-26","2026-02-23"
    }

    _daily_orders = _dd(set)
    for _it in data["order_items"]:
        _d = str(_it.get("CreatedOn",""))[:10]
        _oid = _it.get("OrderId")
        if _oid and _d >= START_DATE:
            _daily_orders[_d].add(_oid)

    _cur = _dtt.strptime(START_DATE, "%Y-%m-%d").date()
    _today = _date.today()
    _gap_dates = []
    while _cur <= _today:
        _day = str(_cur)
        if _day not in KNOWN_CLOSURES and len(_daily_orders.get(_day, set())) < 15:
            _gap_dates.append(_day)
        _cur += _td(days=1)

    if _gap_dates:
        print(f"  Found {len(_gap_dates)} dates to patch: {_gap_dates}")
        _existing_ids = set(i.get("OrderItemId") for i in data["order_items"])
        _total_gap_added = 0
        for _gap_date in _gap_dates:
            _gdt = _dtt.strptime(_gap_date, "%Y-%m-%d")
            _fs = (_gdt - _td(days=2)).strftime("%Y-%m-%d")
            _page = 1
            _tp = 1
            _found = []
            _seen = set()
            while _page <= _tp:
                _ep = f"api/datadump/v2/orderitems/7/{_page}/{_fs}/{LOCATION_ID}"
                try:
                    _res = api_get(_ep)
                    _its = _res.get("data", []) if isinstance(_res, dict) else []
                    _tp = _res.get("pages", 1) if isinstance(_res, dict) else 1
                    for _i in _its:
                        _k = _i.get("OrderItemId")
                        if _k and _k not in _seen:
                            _seen.add(_k)
                            if str(_i.get("CreatedOn","")).startswith(_gap_date):
                                _found.append(_i)
                    _page += 1
                    time.sleep(0.2)
                except Exception as _e:
                    print(f"    Error on {_gap_date}: {_e}")
                    break
            _added = 0
            for _i in _found:
                if _i.get("OrderItemId") not in _existing_ids:
                    data["order_items"].append(_i)
                    _existing_ids.add(_i.get("OrderItemId"))
                    _added += 1
            _total_gap_added += _added
            if _added > 0:
                print(f"  {_gap_date}: recovered {_added} items")
        print(f"  Gap detection complete. Total recovered: {_total_gap_added}")
    else:
        print("  No gaps detected.")

    # 5. Time clocks
    print("\n[5/6] Time clocks...")
    all_clocks = []
    current = datetime.strptime(START_DATE, "%Y-%m-%d")
    end_dt = datetime.strptime(END_DATE, "%Y-%m-%d")
    while current <= end_dt:
        window_end = min(current + timedelta(days=29), end_dt)
        sd = current.strftime("%Y-%m-%dT00:00:00")
        ed = window_end.strftime("%Y-%m-%dT23:59:59")
        try:
            tc = api_post("api/report/timeClocks", {"startDate": sd, "endDate": ed, "pageIndex": 0, "pageSize": 200})
            records = tc if isinstance(tc, list) else tc.get("data", [])
            all_clocks.extend(records)
            print(f"  Clocks {current.strftime('%b')}: {len(records)} records")
        except Exception as e:
            print(f"  Clocks {current.strftime('%b')}: ERROR {e}")
        current = window_end + timedelta(days=1)
        time.sleep(0.3)
    data["time_clocks"] = all_clocks
    print(f"  TOTAL: {len(all_clocks)} clock records")

    # 6. Categories
    print("\n[6/6] Categories...")
    cats = api_get("api/getCategoriesByCompany/1")
    data["categories"] = cats.get("data", []) if isinstance(cats, dict) else cats
    print(f"  {len(data['categories'])} categories")


    # 7. Stock levels by SKU
    print("\n[7/6] Fetching stock levels by SKU...")
    import time as _time
    skus = set()
    for item in data["order_items"]:
        sku = item.get("Sku","").strip()
        if sku and len(sku) > 2:
            skus.add(sku)
    stock_data = {}
    errors = 0
    for i, sku in enumerate(sorted(skus)):
        try:
            r = httpx.get(f"{BASE_URL}/api/getStockByProductSKU/{sku}/{LOCATION_ID}",
                          params={"Token": TOKEN}, timeout=10)
            if r.status_code == 200:
                stock_data[sku] = float(r.text.strip())
            _time.sleep(0.1)
        except:
            errors += 1
    print(f"  Fetched {len(stock_data)} stock levels ({errors} errors)")
    data["stock_levels"] = stock_data
    stock_file = DATA_DIR / "stock_levels.json"
    with open(stock_file, "w") as f:
        json.dump(stock_data, f)
    print(f"  Stock levels saved to {stock_file}")

    # Save cache
    with open(cache_file, "w") as f:
        json.dump(data, f)
    print(f"\nData cached to {cache_file}")

    return data


# ─── Classification Logic ────────────────────────────────────────────────────

def classify_item(name, sku):
    """Classify a line item as grooming service or retail product, with sub-type."""
    name_upper = (name or "").upper()
    sku_str = str(sku or "")

    is_groom = False
    service_type = None
    groom_category = None

    if sku_str.startswith("987") or sku_str.startswith("9870"):
        is_groom = True
        if "FG " in name_upper or "FG/" in name_upper or name_upper.startswith("FG "):
            service_type = "Full Groom"
        elif "MG " in name_upper or "MG/" in name_upper or name_upper.startswith("MG "):
            service_type = "Mini Groom"
        elif "LUX" in name_upper:
            service_type = "Luxury Bath"
        elif "CLSC" in name_upper or "CLASSIC" in name_upper:
            service_type = "Classic Bath"
        else:
            service_type = "Other Groom"
        groom_category = "core"

    elif sku_str.startswith("765"):
        is_groom = True
        groom_category = "spa"
        if "HAIRY BEAST" in name_upper or "DESHED" in name_upper:
            service_type = "Hairy Beast (Deshedding)"
        elif "DIAMOND" in name_upper or "SPOTLESS" in name_upper:
            service_type = "Diamond DG (Spotless and Bright)"
        elif "PLUSH PUP" in name_upper or "SMOOTH" in name_upper:
            service_type = "Plush Pup (Smooth & Silky)"
        elif "TEDDY BEAR" in name_upper or "VOLUME" in name_upper:
            service_type = "The Teddy Bear (Maximum Volume)"
        elif "BLUEBERRY" in name_upper:
            service_type = "Blueberry Pie Facial"
        elif "CREATIVE" in name_upper:
            service_type = "Creative Groom"
        elif "HAPPY" in name_upper or "BARK DAY" in name_upper:
            service_type = "Happy Bark Day"
        else:
            service_type = "SPA Upgrade (Other)"

    elif sku_str.startswith("543"):
        is_groom = True
        groom_category = "addon"
        if "TEETH" in name_upper or "BRUSH" in name_upper:
            service_type = "Teeth Brushing"
        elif "HANDLING" in name_upper or "EXTRA CARE" in name_upper:
            service_type = "Handling / Extra Care"
        elif "DE-MAT" in name_upper or "DEMAT" in name_upper:
            service_type = "De-Matting"
        elif "PAW PAD" in name_upper and "SCISSOR" in name_upper:
            service_type = "Paw Pads & Scissor Feet"
        elif "PAW PAD" in name_upper:
            service_type = "Paw Pads"
        elif "HYPO" in name_upper or "MEDICATED" in name_upper:
            service_type = "Medicated / Hypo Shampoo"
        elif "FLEA" in name_upper or "TICK" in name_upper:
            service_type = "Flea & Tick Shampoo"
        elif "ANAL" in name_upper:
            service_type = "Anal Glands"
        elif "SANITARY" in name_upper:
            service_type = "Sanitary Trim"
        elif "EAR" in name_upper:
            service_type = "Ear Cleaning"
        elif "NAIL" in name_upper and "POLISH" in name_upper:
            service_type = "Nail Polish"
        elif "FACE" in name_upper:
            service_type = "Face Trim"
        elif "SCISSOR" in name_upper:
            service_type = "Scissor Cut"
        elif "BRUSH" in name_upper:
            service_type = "Brushing Out"
        elif "FLAT RATE" in name_upper or "75LBS" in name_upper or "75 LBS" in name_upper:
            service_type = "FG Flat Rate 75lbs+"
        else:
            service_type = "Add-On (Other)"

    elif sku_str.startswith("432"):
        is_groom = True
        groom_category = "walkin"
        if "NAIL" in name_upper:
            service_type = "Nail Trim / Grind"
        elif "DREMEL" in name_upper:
            service_type = "Nail Dremel"
        else:
            service_type = "Walk-In Service"

    elif sku_str.startswith("321"):
        is_groom = True
        groom_category = "fee"
        if "SHAVE" in name_upper:
            service_type = "Shave Down Fee"
        elif "LATE" in name_upper:
            service_type = "Late Pick-Up Fee"
        elif "SKUNK" in name_upper:
            service_type = "Skunk Service"
        else:
            service_type = "Fee (Other)"

    elif sku_str.startswith("876"):
        is_groom = True
        groom_category = "core"
        service_type = "Cat Groom"

    has_retail_sku = len(sku_str) >= 10 and sku_str.isdigit()
    if not is_groom and not has_retail_sku:
        lower = (name or "").lower()
        groom_keywords = ["full groom", "mini groom", "luxury bath", "classic bath",
                          "fg /", "mg /", "lux bath", "clsc bath", "spa upgrade",
                          "add-on", "add on", "walk-in", "nail trim", "teeth brush",
                          "deshed", "hairy beast", "diamond dg", "plush pup",
                          "de-matting", "handling", "paw pad", "anal gland",
                          "blueberry", "creative groom", "teddy bear"]
        for kw in groom_keywords:
            if kw in lower:
                is_groom = True
                if "full groom" in lower or "fg /" in lower or "fg/" in lower:
                    service_type = "Full Groom"
                    groom_category = "core"
                elif "mini groom" in lower or "mg /" in lower or "mg/" in lower:
                    service_type = "Mini Groom"
                    groom_category = "core"
                elif "luxury bath" in lower or "lux bath" in lower:
                    service_type = "Luxury Bath"
                    groom_category = "core"
                elif "classic bath" in lower or "clsc bath" in lower:
                    service_type = "Classic Bath"
                    groom_category = "core"
                else:
                    service_type = service_type or "Service (Misc)"
                    groom_category = groom_category or "addon"
                break

    if not is_groom and ("GIFT CARD" in name_upper or "GIFT CERT" in name_upper):
        return {
            "is_groom": False,
            "is_retail": False,
            "is_gift_card": True,
            "service_type": None,
            "groom_category": None,
            "retail_category": "Gift Cards",
            "dog_size": None,
            "is_doodle": False,
        }

    if "DEPOSIT" in name_upper or "NO-SHOW" in name_upper or "NO SHOW" in name_upper:
        return {
            "is_groom": False,
            "is_retail": False,
            "is_gift_card": False,
            "service_type": "Deposit/No-Show",
            "groom_category": "exclude",
            "retail_category": None,
            "dog_size": None,
            "is_doodle": False,
        }

    dog_size = None
    if is_groom and groom_category == "core":
        if "XXLG" in name_upper or "XX-LG" in name_upper or "XXL" in name_upper:
            dog_size = "Over 100 lbs"
        elif "XLG" in name_upper or "X-LG" in name_upper or " XL " in name_upper or name_upper.endswith(" XL"):
            dog_size = "76-100 lbs"
        elif " LG" in name_upper or "/LG" in name_upper or "/ LG" in name_upper:
            dog_size = "41-75 lbs"
        elif " MD" in name_upper or "/MD" in name_upper or "/ MD" in name_upper or "MED" in name_upper:
            dog_size = "21-40 lbs"
        elif " SM" in name_upper or "/SM" in name_upper or "/ SM" in name_upper:
            dog_size = "0-20 lbs"
        elif "FLAT" in name_upper:
            dog_size = None

    is_doodle = False
    if is_groom:
        if "POODLE" in name_upper or "DOODLE" in name_upper:
            is_doodle = True

    retail_category = None
    if not is_groom:
        lower = (name or "").lower()
        if any(x in lower for x in ["cookie", "bakery bulk", "treat", "chip", "jerky", "pink bag",
                                       "farmers market bulk", "bakery bay bulk"]):
            retail_category = "Treats & Chews"
        elif any(x in lower for x in ["bully", "collagen", "yak", "antler", "marrow", "trachea",
                                       "pig ear", "cow ear", "turkey tendon", "chicken feet",
                                       "chew stick", "natural chew", "rawhide", "bully stick",
                                       "spring", "beef bone"]):
            retail_category = "Natural Chews"
        elif any(x in lower for x in ["toy", "kong", "lamb chop", "squeaker", "ball", "rope",
                                       "plush", "chew toy", "fetch", "tug", "frisbee",
                                       "multipet", "tall tails", "godog", "floppyknots",
                                       "floppy knots", "destroyer", "thoozy", "wunderball"]):
            retail_category = "Toys"
        elif any(x in lower for x in ["birthday", "cupcake", "gelato", "cannoli", "cake",
                                       "preppy puppy", "lucky b", "donut", "holiday", "raffle"]):
            retail_category = "Bakery & Birthday"
        elif any(x in lower for x in ["dry food", "kibble", "farmina", "n&d"]):
            retail_category = "Dry Food"
        elif any(x in lower for x in ["wet food", "canned", "pate", "stew"]):
            retail_category = "Wet Food"
        elif any(x in lower for x in ["supplement", "probiotic", "hip & joint", "hip and joint",
                                       "skin & coat", "digestive", "functional", "vitamin",
                                       "plaqueoff", "calming"]):
            retail_category = "Supplements & Health"
        elif any(x in lower for x in ["shampoo", "conditioner", "spray", "cologne", "wipe",
                                       "grooming", "brush", "comb", "nail clipper"]):
            retail_category = "Grooming Supplies"
        elif any(x in lower for x in ["collar", "leash", "harness", "bandana", "bow tie",
                                       "apparel", "jacket", "sweater", "hat", "birdie dawg"]):
            retail_category = "Apparel & Lifestyle"
        elif any(x in lower for x in ["bowl", "feeder", "bed", "crate", "carrier", "mat"]):
            retail_category = "Accessories"
        elif any(x in lower for x in ["topper", "mix-in", "broth", "goat milk", "pumpkin puree"]):
            retail_category = "Toppers & Mix-ins"
        elif any(x in lower for x in ["raffle", "donation", "event"]):
            retail_category = "Other"
        elif "miscellaneous" in lower or sku_str == "000":
            retail_category = "Other"
        else:
            retail_category = "Other"

    return {
        "is_groom": is_groom,
        "is_retail": not is_groom,
        "is_gift_card": False,
        "service_type": service_type,
        "groom_category": groom_category,
        "retail_category": retail_category,
        "dog_size": dog_size,
        "is_doodle": is_doodle,
    }


# ─── Data Transformation ─────────────────────────────────────────────────────

def transform_data(raw):
    """Transform raw API data into analysis DataFrames."""
    print("\n" + "=" * 60)
    print("TRANSFORMING DATA")
    print("=" * 60)

    emp_map = {}
    for e in raw.get("employees", []):
        eid = e.get("Id")
        name = f"{e.get('FirstName', '')} {e.get('LastName', '')}".strip()
        emp_map[eid] = name

    items = raw.get("order_items", [])
    print(f"Processing {len(items)} line items...")

    rows = []
    for item in items:
        name = item.get("Name", "")
        sku = item.get("Sku", "")
        cls = classify_item(name, sku)

        net_sales = float(item.get("Price", 0)) * float(item.get("Quantity", 1)) - float(item.get("Discount", 0))

        rows.append({
            "order_id": item.get("OrderId"),
            "item_id": item.get("OrderItemId"),
            "customer_id": item.get("CustomerId"),
            "created": item.get("CreatedOn"),
            "name": name,
            "sku": sku,
            "price": float(item.get("Price", 0)),
            "quantity": float(item.get("Quantity", 1)),
            "discount": float(item.get("Discount", 0)),
            "net_sales": net_sales,
            "cost": float(item.get("Cost", 0)),
            "salesperson": item.get("SalesPerson", ""),
            **cls,
        })

    df_items = pd.DataFrame(rows)
    df_items["created"] = pd.to_datetime(df_items["created"], utc=True).dt.tz_convert("America/New_York")
    df_items["month"] = df_items["created"].dt.month
    df_items["month_name"] = df_items["created"].dt.strftime("%b")
    df_items["day_of_week"] = df_items["created"].dt.day_name()

    orders = raw.get("orders", [])
    print(f"Processing {len(orders)} orders...")

    order_rows = []
    for o in orders:
        order_rows.append({
            "order_id": o.get("OrderId"),
            "customer_id": o.get("CustomerId"),
            "employee_id": o.get("EmployeeId"),
            "created": o.get("CreatedOn"),
            "subtotal": float(o.get("SubTotal", 0)),
            "discount_total": float(o.get("DiscountTotal", 0)),
            "tax_total": float(o.get("TaxTotal", 0)),
            "tips": float(o.get("Tips", 0)),
            "total": float(o.get("Total", 0)),
        })

    if order_rows:
        df_orders = pd.DataFrame(order_rows)
        df_orders["created"] = pd.to_datetime(df_orders["created"], utc=True).dt.tz_convert("America/New_York")
        df_orders["month"] = df_orders["created"].dt.month
    else:
        print("  No orders from API — reconstructing from line items...")
        grp = df_items.groupby("order_id").agg(
            customer_id=("customer_id", "first"),
            created=("created", "first"),
            subtotal=("net_sales", "sum"),
            total=("net_sales", "sum"),
        ).reset_index()
        grp["employee_id"] = None
        grp["discount_total"] = 0.0
        grp["tax_total"] = 0.0
        grp["tips"] = 0.0
        df_orders = grp
        df_orders["month"] = df_orders["created"].dt.month
        print(f"  Reconstructed {len(df_orders)} orders from line items")

    clocks = raw.get("time_clocks", [])
    clock_rows = []
    for c in clocks:
        if isinstance(c, dict):
            clock_rows.append({
                "employee_id": c.get("EmployeeId"),
                "employee_name": c.get("EmployeeName", ""),
                "clock_in": c.get("ClockIn"),
                "clock_out": c.get("ClockOut"),
                "hours": float(c.get("TotalHours", 0) or 0),
            })
    df_clocks = pd.DataFrame(clock_rows) if clock_rows else pd.DataFrame(columns=["employee_id", "employee_name", "hours"])

    print(f"Items: {len(df_items)}, Orders: {len(df_orders)}, Clock records: {len(df_clocks)}")

    df_valid = df_items[df_items["groom_category"] != "exclude"].copy()

    return {
        "df_items": df_valid,
        "df_orders": df_orders,
        "df_clocks": df_clocks,
        "emp_map": emp_map,
        "all_items": df_items,
    }


# ─── Analysis Functions ───────────────────────────────────────────────────────

def compute_executive_summary(df, df_orders):
    groom = df[df["is_groom"] == True]
    retail = df[df["is_retail"] == True]
    gifts = df[df["is_gift_card"] == True]

    total_net = df["net_sales"].sum()
    total_gross = (df["price"] * df["quantity"]).sum()
    total_returns = df[df["net_sales"] < 0]["net_sales"].sum()
    total_discounts = df["discount"].sum()
    total_units = int(df["quantity"].sum())
    total_txns = df_orders.shape[0]
    unique_skus = df["sku"].nunique()
    avg_price = total_gross / total_units if total_units else 0
    avg_txn = total_net / total_txns if total_txns else 0

    return {
        "total_net_sales": total_net,
        "total_gross_sales": total_gross,
        "total_returns": total_returns,
        "total_discounts": total_discounts,
        "total_units": total_units,
        "total_transactions": total_txns,
        "total_skus": unique_skus,
        "avg_price_per_item": avg_price,
        "avg_transaction_value": avg_txn,
        "grooming_revenue": groom["net_sales"].sum(),
        "retail_revenue": retail["net_sales"].sum(),
        "gift_card_revenue": gifts["net_sales"].sum(),
        "grooming_pct": groom["net_sales"].sum() / total_net * 100 if total_net else 0,
        "retail_pct": retail["net_sales"].sum() / total_net * 100 if total_net else 0,
    }


def compute_service_mix(df):
    groom = df[df["is_groom"] == True].copy()
    if groom.empty:
        return pd.DataFrame()

    mix = groom.groupby("service_type").agg(
        units=("quantity", "sum"),
        revenue=("net_sales", "sum"),
        sku_count=("sku", "nunique"),
    ).reset_index()
    mix["pct_units"] = mix["units"] / mix["units"].sum() * 100
    mix["pct_revenue"] = mix["revenue"] / mix["revenue"].sum() * 100
    mix["avg_ticket"] = mix["revenue"] / mix["units"]
    mix = mix.sort_values("revenue", ascending=False)
    return mix


def compute_dog_sizes(df):
    core = df[(df["groom_category"] == "core") & (df["dog_size"].notna())].copy()
    if core.empty:
        return {}, {}, {}

    all_sizes = core.groupby("dog_size").agg(
        units=("quantity", "sum"),
        revenue=("net_sales", "sum"),
    ).reset_index()
    all_sizes["pct_units"] = all_sizes["units"] / all_sizes["units"].sum() * 100
    all_sizes["pct_revenue"] = all_sizes["revenue"] / all_sizes["revenue"].sum() * 100
    all_sizes["avg_ticket"] = all_sizes["revenue"] / all_sizes["units"]

    std = core[core["is_doodle"] == False]
    dood = core[core["is_doodle"] == True]

    std_sizes = std.groupby("dog_size").agg(
        units=("quantity", "sum"), revenue=("net_sales", "sum")
    ).reset_index()
    std_sizes["avg_ticket"] = std_sizes["revenue"] / std_sizes["units"]

    dood_sizes = dood.groupby("dog_size").agg(
        units=("quantity", "sum"), revenue=("net_sales", "sum")
    ).reset_index()
    dood_sizes["avg_ticket"] = dood_sizes["revenue"] / dood_sizes["units"]

    size_order = ["0-20 lbs", "21-40 lbs", "41-75 lbs", "76-100 lbs", "Over 100 lbs"]
    for frame in [all_sizes, std_sizes, dood_sizes]:
        frame["_order"] = frame["dog_size"].map({s: i for i, s in enumerate(size_order)})
        frame.sort_values("_order", inplace=True)
        frame.drop("_order", axis=1, inplace=True)

    return all_sizes, std_sizes, dood_sizes


def compute_monthly_performance(df, df_orders):
    groom = df[df["is_groom"] == True]
    retail = df[df["is_retail"] == True]

    df["_ym"] = df["created"].dt.to_period("M")
    if "created" in df_orders.columns:
        df_orders["_ym"] = df_orders["created"].dt.to_period("M")
    groom["_ym"] = groom["created"].dt.to_period("M")
    retail["_ym"] = retail["created"].dt.to_period("M")

    all_periods = sorted(df["_ym"].dropna().unique())

    months = []
    for p in all_periods:
        m_groom = groom[groom["_ym"] == p]
        m_retail = retail[retail["_ym"] == p]
        m_orders = df_orders[df_orders["_ym"] == p] if "_ym" in df_orders.columns else pd.DataFrame()

        groom_rev = m_groom["net_sales"].sum()
        retail_rev = m_retail["net_sales"].sum()
        total_rev = groom_rev + retail_rev
        tickets = m_orders.shape[0]

        months.append({
            "month": p.month,
            "month_name": p.strftime("%b %Y") if len(all_periods) > 12 or p.year != all_periods[0].year else p.strftime("%b"),
            "net_revenue": total_rev,
            "grooming_rev": groom_rev,
            "retail_rev": retail_rev,
            "groom_pct": groom_rev / total_rev * 100 if total_rev else 0,
            "retail_pct": retail_rev / total_rev * 100 if total_rev else 0,
            "tickets": tickets,
            "avg_ticket": total_rev / tickets if tickets else 0,
        })

    df.drop("_ym", axis=1, inplace=True, errors="ignore")
    df_orders.drop("_ym", axis=1, inplace=True, errors="ignore")

    df_monthly = pd.DataFrame(months)
    df_monthly["mom_growth"] = df_monthly["net_revenue"].pct_change() * 100
    return df_monthly


def compute_top_retail(df, n=50):
    retail = df[df["is_retail"] == True].copy()
    if retail.empty:
        return pd.DataFrame(), pd.DataFrame()

    total_retail = retail["net_sales"].sum()

    by_rev = retail.groupby(["sku", "name", "retail_category"]).agg(
        units=("quantity", "sum"),
        net_sales=("net_sales", "sum"),
    ).reset_index()
    by_rev["pct_retail"] = by_rev["net_sales"] / total_retail * 100
    by_rev["avg_price"] = by_rev["net_sales"] / by_rev["units"]

    top_by_rev = by_rev.sort_values("net_sales", ascending=False).head(n).reset_index(drop=True)
    top_by_rev.index = top_by_rev.index + 1

    top_by_units = by_rev.sort_values("units", ascending=False).head(n).reset_index(drop=True)
    top_by_units.index = top_by_units.index + 1

    return top_by_rev, top_by_units


def compute_grooming_skus(df):
    groom = df[df["is_groom"] == True].copy()
    if groom.empty:
        return pd.DataFrame()

    by_sku = groom.groupby(["sku", "name", "service_type"]).agg(
        units=("quantity", "sum"),
        net_sales=("net_sales", "sum"),
    ).reset_index()
    total_groom = groom["net_sales"].sum()
    by_sku["pct_mix"] = by_sku["net_sales"] / total_groom * 100
    by_sku["avg_ticket"] = by_sku["net_sales"] / by_sku["units"]
    by_sku = by_sku.sort_values("net_sales", ascending=False).reset_index(drop=True)
    by_sku.index = by_sku.index + 1
    return by_sku


def compute_category_top10s(df):
    retail = df[df["is_retail"] == True].copy()
    if retail.empty:
        return {}

    results = {}
    for cat in retail["retail_category"].unique():
        if cat in ("Other", None):
            continue
        cat_df = retail[retail["retail_category"] == cat]
        cat_total = cat_df["net_sales"].sum()

        by_product = cat_df.groupby(["name"]).agg(
            units=("quantity", "sum"),
            revenue=("net_sales", "sum"),
        ).reset_index()

        top_rev = by_product.sort_values("revenue", ascending=False).head(10)
        top_units = by_product.sort_values("units", ascending=False).head(10)

        results[cat] = {
            "total": cat_total,
            "top_by_revenue": top_rev,
            "top_by_units": top_units,
        }

    return results


def compute_brand_breakdown(df):
    retail = df[df["is_retail"] == True].copy()
    if retail.empty:
        return {}

    def extract_brand(name):
        n = (name or "").strip()
        brands = [
            ("WGB ", "WGB / Woof Gang"), ("Woof Gang", "WGB / Woof Gang"),
            ("Farmina", "Farmina"), ("KONG ", "KONG"), ("Kong ", "KONG"),
            ("Tall Tails", "Tall Tails"), ("Multipet", "Multipet"),
            ("BIXBI", "BIXBI"), ("Vital Essentials", "Vital Essentials"),
            ("Himalayan", "Himalayan"), ("H&K", "H&K / Haute Diggity Dog"),
            ("Haute Diggity", "H&K / Haute Diggity Dog"),
            ("PlaqueOff", "PlaqueOff"), ("Nootie", "Nootie"),
            ("goDog", "goDog"), ("Pets First", "Pets First"),
            ("Earth Animal", "Earth Animal"), ("Grandma Lucy", "Grandma Lucy's"),
            ("Puppy Cake", "Puppy Cake"), ("Swell Gelato", "Swell Gelato"),
            ("Lucky B", "Lucky B"), ("Preppy Puppy", "Preppy Puppy"),
            ("Birdie Dawg", "Birdie Dawg"),
            ("Cookie", "Store/Bakery Items"), ("Bakery Bulk", "Store/Bakery Items"),
            ("Mini Birthday", "Store/Bakery Items"), ("Birthday Cupcake", "Store/Bakery Items"),
            ("Mini PB", "Store/Bakery Items"),
            ("Wunderball", "Wunderball"),
        ]
        for prefix, brand in brands:
            if prefix.lower() in n.lower():
                return brand
        if n.startswith("Miscellaneous"):
            return "Other/Unbranded"
        return "Other/Unbranded"

    retail["brand"] = retail["name"].apply(extract_brand)

    results = {}
    for cat in retail["retail_category"].unique():
        if cat in ("Other", None):
            continue
        cat_df = retail[retail["retail_category"] == cat]
        cat_total = cat_df["net_sales"].sum()

        by_brand = cat_df.groupby("brand").agg(
            net_sales=("net_sales", "sum"),
            units=("quantity", "sum"),
            sku_count=("sku", "nunique"),
        ).reset_index()
        by_brand["pct_category"] = by_brand["net_sales"] / cat_total * 100
        by_brand["avg_price"] = by_brand["net_sales"] / by_brand["units"]
        by_brand["revenue_per_sku"] = by_brand["net_sales"] / by_brand["sku_count"]
        by_brand = by_brand.sort_values("net_sales", ascending=False).head(10)

        results[cat] = {"total": cat_total, "brands": by_brand}

    return results


def compute_top_customers(df, df_orders):
    order_summary = df_orders.groupby("customer_id").agg(
        total_spend=("subtotal", "sum"),
        transactions=("order_id", "count"),
        visit_days=("created", lambda x: x.dt.date.nunique()),
    ).reset_index()
    order_summary["spend_per_visit"] = order_summary["total_spend"] / order_summary["visit_days"]
    order_summary["avg_txn"] = order_summary["total_spend"] / order_summary["transactions"]

    def tier(row):
        spv = row["spend_per_visit"]
        if spv >= 150: return "Premium"
        elif spv >= 100: return "High"
        elif spv >= 60: return "Standard"
        return "Value"
    order_summary["tier"] = order_summary.apply(tier, axis=1)

    top = order_summary.sort_values("total_spend", ascending=False).head(50)
    return top


def compute_customer_intelligence(df, df_orders):
    order_items_agg = df.groupby("order_id").agg(
        has_groom=("is_groom", "any"),
        has_retail=("is_retail", "any"),
        has_gift=("is_gift_card", "any"),
        total=("net_sales", "sum"),
    ).reset_index()

    def ticket_type(row):
        if row["has_groom"] and row["has_retail"]: return "Hybrid (Groom + Retail)"
        elif row["has_groom"]: return "Grooming Only"
        elif row["has_retail"]: return "Retail Only"
        elif row["has_gift"]: return "Gift Card Only"
        return "Other"

    order_items_agg["ticket_type"] = order_items_agg.apply(ticket_type, axis=1)

    ticket_comp = order_items_agg.groupby("ticket_type").agg(
        tickets=("order_id", "count"),
        revenue=("total", "sum"),
    ).reset_index()
    total_tickets = ticket_comp["tickets"].sum()
    total_rev = ticket_comp["revenue"].sum()
    ticket_comp["pct_tickets"] = ticket_comp["tickets"] / total_tickets * 100
    ticket_comp["pct_revenue"] = ticket_comp["revenue"] / total_rev * 100
    ticket_comp["avg_ticket"] = ticket_comp["revenue"] / ticket_comp["tickets"]

    named_orders = df_orders[df_orders["customer_id"] > 0]
    cust_visits = named_orders.groupby("customer_id").agg(
        visits=("created", lambda x: x.dt.date.nunique()),
        total_spend=("subtotal", "sum"),
    ).reset_index()

    def freq_bucket(v):
        if v == 1: return "1 visit (trial)"
        elif v == 2: return "2 visits"
        elif v <= 4: return "3-4 visits (developing)"
        elif v <= 8: return "5-8 visits (loyal)"
        elif v <= 12: return "9-12 visits (committed)"
        return "13+ visits (champions)"

    cust_visits["bucket"] = cust_visits["visits"].apply(freq_bucket)
    freq = cust_visits.groupby("bucket").agg(
        customers=("customer_id", "count"),
        revenue=("total_spend", "sum"),
    ).reset_index()
    freq["pct_customers"] = freq["customers"] / freq["customers"].sum() * 100
    freq["pct_revenue"] = freq["revenue"] / freq["revenue"].sum() * 100
    freq["avg_spend"] = freq["revenue"] / freq["customers"]

    end_date = pd.Timestamp.now().normalize()
    last_visit = named_orders.groupby("customer_id")["created"].max().reset_index()
    last_visit["created"] = last_visit["created"].dt.tz_localize(None) if last_visit["created"].dt.tz is not None else last_visit["created"]
    last_visit["days_since"] = (end_date - last_visit["created"]).dt.days
    cust_spend = named_orders.groupby("customer_id")["subtotal"].sum().reset_index()
    last_visit = last_visit.merge(cust_spend, on="customer_id")

    def churn_segment(days):
        if days <= 30: return "Active (last 30 days)"
        elif days <= 60: return "Recent (31-60 days)"
        elif days <= 90: return "At Risk (61-90 days)"
        elif days <= 180: return "Lapsing (91-180 days)"
        return "Lost (180+ days)"

    last_visit["segment"] = last_visit["days_since"].apply(churn_segment)
    churn = last_visit.groupby("segment").agg(
        customers=("customer_id", "count"),
        total_spend=("subtotal", "sum"),
    ).reset_index()
    churn["pct"] = churn["customers"] / churn["customers"].sum() * 100
    churn["avg_spend"] = churn["total_spend"] / churn["customers"]

    dow_items = df.copy()
    dow_items["dow"] = dow_items["created"].dt.day_name()
    dow_orders = df_orders.copy()
    dow_orders["dow"] = dow_orders["created"].dt.day_name()

    dow_rev = dow_items.groupby("dow")["net_sales"].sum()
    dow_groom_rev = dow_items[dow_items["is_groom"]].groupby("dow")["net_sales"].sum()
    dow_retail_rev = dow_items[dow_items["is_retail"]].groupby("dow")["net_sales"].sum()
    dow_tickets = dow_orders.groupby("dow")["order_id"].count()

    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    dow_df = pd.DataFrame({"day": day_order})
    dow_df["revenue"] = dow_df["day"].map(dow_rev).fillna(0)
    dow_df["grooming_rev"] = dow_df["day"].map(dow_groom_rev).fillna(0)
    dow_df["retail_rev"] = dow_df["day"].map(dow_retail_rev).fillna(0)
    dow_df["tickets"] = dow_df["day"].map(dow_tickets).fillna(0).astype(int)
    total_week_rev = dow_df["revenue"].sum()
    dow_df["pct_week"] = dow_df["revenue"] / total_week_rev * 100
    dow_df["avg_ticket"] = dow_df["revenue"] / dow_df["tickets"].replace(0, 1)
    dow_df["retail_pct"] = dow_df["retail_rev"] / dow_df["revenue"].replace(0, 1) * 100

    return {
        "ticket_composition": ticket_comp,
        "visit_frequency": freq,
        "churn_risk": churn,
        "day_of_week": dow_df,
    }


def compute_team_performance(df, df_orders, df_clocks):
    groom = df[df["is_groom"] == True].copy()

    team = groom.groupby("salesperson").agg(
        revenue=("net_sales", "sum"),
        appointments=("order_id", "nunique"),
        items=("quantity", "sum"),
    ).reset_index()
    team = team[team["salesperson"] != ""].copy()
    team["per_appt"] = team["revenue"] / team["appointments"]
    team = team.sort_values("revenue", ascending=False)

    core = groom[groom["groom_category"] == "core"]
    spa = groom[groom["groom_category"] == "spa"]
    addon = groom[groom["groom_category"] == "addon"]

    core_by_groomer = core.groupby("salesperson")["order_id"].nunique().reset_index()
    core_by_groomer.columns = ["salesperson", "core_visits"]

    spa_by_groomer = spa.groupby("salesperson")["order_id"].nunique().reset_index()
    spa_by_groomer.columns = ["salesperson", "spa_orders"]

    addon_by_groomer = addon.groupby("salesperson")["order_id"].nunique().reset_index()
    addon_by_groomer.columns = ["salesperson", "addon_orders"]

    teeth = groom[groom["service_type"] == "Teeth Brushing"]
    teeth_by_groomer = teeth.groupby("salesperson")["order_id"].nunique().reset_index()
    teeth_by_groomer.columns = ["salesperson", "teeth_orders"]

    team = team.merge(core_by_groomer, on="salesperson", how="left")
    team = team.merge(spa_by_groomer, on="salesperson", how="left")
    team = team.merge(addon_by_groomer, on="salesperson", how="left")
    team = team.merge(teeth_by_groomer, on="salesperson", how="left")

    team["spa_orders"] = team["spa_orders"].fillna(0)
    team["addon_orders"] = team["addon_orders"].fillna(0)
    team["teeth_orders"] = team["teeth_orders"].fillna(0)
    team["core_visits"] = team["core_visits"].fillna(0)

    team["spa_pct"] = (team["spa_orders"] / team["core_visits"].replace(0, 1) * 100)
    team["addon_pct"] = (team["addon_orders"] / team["core_visits"].replace(0, 1) * 100)
    team["teeth_pct"] = (team["teeth_orders"] / team["core_visits"].replace(0, 1) * 100)

    if not df_clocks.empty:
        hours_by_emp = df_clocks.groupby("employee_name")["hours"].sum().reset_index()
        team = team.merge(hours_by_emp, left_on="salesperson", right_on="employee_name", how="left")
        team["per_hour"] = team["revenue"] / team["hours"].replace(0, float("nan"))
    else:
        team["hours"] = float("nan")
        team["per_hour"] = float("nan")

    return team


# ─── Excel Workbook Generator ────────────────────────────────────────────────

def _style_header(ws, row, max_col, fill_color=DARK_TEAL):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="Inter", size=10, bold=True, color=WHITE)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_title(ws, row, text, col=1, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Inter", size=14, bold=True, color=PAW_MAGENTA)
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)


def _style_subtitle(ws, row, text, col=1, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Inter", size=10, italic=True, color="666666")
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)


def _fmt_currency(val):
    if pd.isna(val): return ""
    return f"${val:,.2f}"


def _fmt_pct(val):
    if pd.isna(val): return ""
    return f"{val:.1f}%"


def _fmt_int(val):
    if pd.isna(val): return ""
    return f"{int(val):,}"


def _auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _zebra_rows(ws, start_row, end_row, max_col):
    light = PatternFill(start_color=LIGHT_PINK, end_color=LIGHT_PINK, fill_type="solid")
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = light


def write_executive_summary(wb, summary, df_items):
    ws = wb.create_sheet("Executive Summary")
    _style_title(ws, 1, STORE_NAME, merge_to=5)
    _style_subtitle(ws, 2, f"Performance Analysis  |  {START_DATE} to {END_DATE}", merge_to=5)

    row = 4
    ws.cell(row=row, column=1, value="Store Overview").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 1
    for i, h in enumerate(["Metric", "Value"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 2)

    metrics = [
        ("Total Net Sales", _fmt_currency(summary["total_net_sales"])),
        ("Total Gross Sales", _fmt_currency(summary["total_gross_sales"])),
        ("Total Returns", _fmt_currency(summary["total_returns"])),
        ("Total Discounts", _fmt_currency(summary["total_discounts"])),
        ("Total Units Sold", _fmt_int(summary["total_units"])),
        ("Total Transactions", _fmt_int(summary["total_transactions"])),
        ("Total SKUs", _fmt_int(summary["total_skus"])),
        ("Avg Price per Item", _fmt_currency(summary["avg_price_per_item"])),
        ("Average Transaction Value", _fmt_currency(summary["avg_transaction_value"])),
        ("Date Range", f"{START_DATE} to {END_DATE}"),
    ]
    for i, (metric, val) in enumerate(metrics):
        ws.cell(row=row + 1 + i, column=1, value=metric).font = Font(name="Inter", size=10)
        ws.cell(row=row + 1 + i, column=2, value=val).font = Font(name="Inter", size=10, bold=True)
    _zebra_rows(ws, row + 1, row + len(metrics), 2)

    row += len(metrics) + 2
    ws.cell(row=row, column=1, value="Revenue Mix").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 1
    for i, h in enumerate(["Category", "Revenue", "% of Total"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 3)

    mix_data = [
        ("Grooming Services", summary["grooming_revenue"], summary["grooming_pct"]),
        ("Retail Sales", summary["retail_revenue"], summary["retail_pct"]),
        ("Gift Cards", summary["gift_card_revenue"],
         summary["gift_card_revenue"] / summary["total_net_sales"] * 100 if summary["total_net_sales"] else 0),
    ]
    for i, (cat, rev, pct) in enumerate(mix_data):
        ws.cell(row=row + 1 + i, column=1, value=cat)
        ws.cell(row=row + 1 + i, column=2, value=_fmt_currency(rev))
        ws.cell(row=row + 1 + i, column=3, value=_fmt_pct(pct))
    _zebra_rows(ws, row + 1, row + len(mix_data), 3)

    groom = df_items[df_items["is_groom"] == True]
    if not groom.empty:
        row += len(mix_data) + 2
        ws.cell(row=row, column=1, value="SKU Concentration -- Grooming Services").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
        _style_subtitle(ws, row + 1, "Pareto/ABC analysis of grooming service SKUs only", merge_to=5)
        row += 2
        for i, h in enumerate(["Class", "SKU Count", "% of SKUs", "Revenue", "% of Revenue"], 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, 5)

        sku_rev = groom.groupby("sku")["net_sales"].sum().sort_values(ascending=False)
        total_groom_rev = sku_rev.sum()
        cumsum = sku_rev.cumsum() / total_groom_rev
        a_count = (cumsum <= 0.80).sum()
        b_count = ((cumsum > 0.80) & (cumsum <= 0.95)).sum()
        c_count = (cumsum > 0.95).sum()
        total_skus = len(sku_rev)

        abc = [
            ("A-Class (Top 80%)", a_count, a_count/total_skus*100, sku_rev.iloc[:a_count].sum(), sku_rev.iloc[:a_count].sum()/total_groom_rev*100),
            ("B-Class (Next 15%)", b_count, b_count/total_skus*100, sku_rev.iloc[a_count:a_count+b_count].sum(), sku_rev.iloc[a_count:a_count+b_count].sum()/total_groom_rev*100),
            ("C-Class (Remaining 5%)", c_count, c_count/total_skus*100, sku_rev.iloc[a_count+b_count:].sum(), sku_rev.iloc[a_count+b_count:].sum()/total_groom_rev*100),
        ]
        for i, (cls, cnt, pct_s, rev, pct_r) in enumerate(abc):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=cls)
            ws.cell(row=r, column=2, value=cnt)
            ws.cell(row=r, column=3, value=_fmt_pct(pct_s))
            ws.cell(row=r, column=4, value=_fmt_currency(rev))
            ws.cell(row=r, column=5, value=_fmt_pct(pct_r))
        _zebra_rows(ws, row + 1, row + 3, 5)

    retail = df_items[df_items["is_retail"] == True]
    if not retail.empty:
        row += 5
        ws.cell(row=row, column=1, value="SKU Concentration -- Retail").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
        _style_subtitle(ws, row + 1, "Pareto/ABC analysis of retail product SKUs only", merge_to=5)
        row += 2
        for i, h in enumerate(["Class", "SKU Count", "% of SKUs", "Revenue", "% of Revenue"], 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, 5)

        sku_rev = retail.groupby("sku")["net_sales"].sum().sort_values(ascending=False)
        total_retail_rev = sku_rev.sum()
        cumsum = sku_rev.cumsum() / total_retail_rev
        a_count = max((cumsum <= 0.80).sum(), 1)
        b_count = max(((cumsum > 0.80) & (cumsum <= 0.95)).sum(), 0)
        c_count = max((cumsum > 0.95).sum(), 0)
        total_skus = len(sku_rev)

        abc = [
            ("A-Class (Top 80%)", a_count, a_count/total_skus*100 if total_skus else 0,
             sku_rev.iloc[:a_count].sum(), sku_rev.iloc[:a_count].sum()/total_retail_rev*100 if total_retail_rev else 0),
            ("B-Class (Next 15%)", b_count, b_count/total_skus*100 if total_skus else 0,
             sku_rev.iloc[a_count:a_count+b_count].sum(),
             sku_rev.iloc[a_count:a_count+b_count].sum()/total_retail_rev*100 if total_retail_rev else 0),
            ("C-Class (Remaining 5%)", c_count, c_count/total_skus*100 if total_skus else 0,
             sku_rev.iloc[a_count+b_count:].sum(),
             sku_rev.iloc[a_count+b_count:].sum()/total_retail_rev*100 if total_retail_rev else 0),
        ]
        for i, (cls, cnt, pct_s, rev, pct_r) in enumerate(abc):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=cls)
            ws.cell(row=r, column=2, value=cnt)
            ws.cell(row=r, column=3, value=_fmt_pct(pct_s))
            ws.cell(row=r, column=4, value=_fmt_currency(rev))
            ws.cell(row=r, column=5, value=_fmt_pct(pct_r))
        _zebra_rows(ws, row + 1, row + 3, 5)

    _auto_width(ws)
    return ws


def write_service_mix(wb, mix_df):
    ws = wb.create_sheet("Service Type Mix")
    _style_title(ws, 1, "Service Type Mix", merge_to=7)
    _style_subtitle(ws, 2, "Grooming service breakdown using actual location service names", merge_to=7)

    row = 4
    headers = ["Service Type", "Units", "% of Units", "Revenue", "% of Revenue", "Avg Ticket", "SKU Count"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 7)

    for i, (_, r) in enumerate(mix_df.iterrows()):
        row += 1
        ws.cell(row=row, column=1, value=r["service_type"])
        ws.cell(row=row, column=2, value=_fmt_int(r["units"]))
        ws.cell(row=row, column=3, value=_fmt_pct(r["pct_units"]))
        ws.cell(row=row, column=4, value=_fmt_currency(r["revenue"]))
        ws.cell(row=row, column=5, value=_fmt_pct(r["pct_revenue"]))
        ws.cell(row=row, column=6, value=_fmt_currency(r["avg_ticket"]))
        ws.cell(row=row, column=7, value=int(r["sku_count"]))

    _zebra_rows(ws, 5, row, 7)

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=_fmt_int(mix_df["units"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=3, value="100.0%").font = Font(bold=True)
    ws.cell(row=row, column=4, value=_fmt_currency(mix_df["revenue"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=5, value="100.0%").font = Font(bold=True)
    ws.cell(row=row, column=6, value=_fmt_currency(mix_df["revenue"].sum() / mix_df["units"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=7, value=int(mix_df["sku_count"].sum())).font = Font(bold=True)

    _auto_width(ws)
    return ws


def write_dog_sizes(wb, all_sizes, std_sizes, dood_sizes):
    ws = wb.create_sheet("Dog Size Breakdown")
    _style_title(ws, 1, "All Dogs -- Size Distribution", merge_to=6)
    _style_subtitle(ws, 2, "Core grooms (Full Groom + Mini Groom + Luxury Bath + Classic Bath)", merge_to=6)

    def write_size_table(ws, start_row, title, data):
        ws.cell(row=start_row, column=1, value=title).font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        r = start_row + 1
        for i, h in enumerate(["Size Bracket", "Units", "% of Units", "Revenue", "% of Revenue", "Avg Ticket"], 1):
            ws.cell(row=r, column=i, value=h)
        _style_header(ws, r, 6)

        for idx, row_data in data.iterrows():
            r += 1
            ws.cell(row=r, column=1, value=row_data["dog_size"])
            ws.cell(row=r, column=2, value=_fmt_int(row_data["units"]))
            ws.cell(row=r, column=3, value=_fmt_pct(row_data.get("pct_units", row_data["units"]/data["units"].sum()*100)))
            ws.cell(row=r, column=4, value=_fmt_currency(row_data["revenue"]))
            ws.cell(row=r, column=5, value=_fmt_pct(row_data.get("pct_revenue", row_data["revenue"]/data["revenue"].sum()*100)))
            ws.cell(row=r, column=6, value=_fmt_currency(row_data["avg_ticket"]))

        _zebra_rows(ws, start_row + 2, r, 6)
        r += 1
        ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=r, column=2, value=_fmt_int(data["units"].sum())).font = Font(bold=True)
        ws.cell(row=r, column=4, value=_fmt_currency(data["revenue"].sum())).font = Font(bold=True)
        ws.cell(row=r, column=6, value=_fmt_currency(data["revenue"].sum()/data["units"].sum() if data["units"].sum() else 0)).font = Font(bold=True)
        return r + 2

    next_row = 4
    if isinstance(all_sizes, pd.DataFrame) and not all_sizes.empty:
        next_row = write_size_table(ws, next_row, "All Dogs", all_sizes)
    if isinstance(std_sizes, pd.DataFrame) and not std_sizes.empty:
        next_row = write_size_table(ws, next_row, "Standard Dogs (Non-Doodle/Poodle)", std_sizes)
    if isinstance(dood_sizes, pd.DataFrame) and not dood_sizes.empty:
        next_row = write_size_table(ws, next_row, "Doodle Dogs (Poodle & Doodle breeds)", dood_sizes)

    if isinstance(std_sizes, pd.DataFrame) and not std_sizes.empty and isinstance(dood_sizes, pd.DataFrame) and not dood_sizes.empty:
        ws.cell(row=next_row, column=1, value="Standard vs Doodle Price Comparison").font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        r = next_row + 1
        for i, h in enumerate(["Size", "Std Avg Ticket", "Doodle Avg Ticket", "Premium $", "Premium %"], 1):
            ws.cell(row=r, column=i, value=h)
        _style_header(ws, r, 5)

        for size in ["0-20 lbs", "21-40 lbs", "41-75 lbs", "76-100 lbs", "Over 100 lbs"]:
            std_row = std_sizes[std_sizes["dog_size"] == size]
            dood_row = dood_sizes[dood_sizes["dog_size"] == size]
            if std_row.empty or dood_row.empty:
                continue
            std_avg = std_row["avg_ticket"].values[0]
            dood_avg = dood_row["avg_ticket"].values[0]
            premium = dood_avg - std_avg
            prem_pct = premium / std_avg * 100 if std_avg else 0
            r += 1
            ws.cell(row=r, column=1, value=size)
            ws.cell(row=r, column=2, value=_fmt_currency(std_avg))
            ws.cell(row=r, column=3, value=_fmt_currency(dood_avg))
            ws.cell(row=r, column=4, value=_fmt_currency(premium))
            ws.cell(row=r, column=5, value=_fmt_pct(prem_pct))

    _auto_width(ws)
    return ws


def write_grooming_skus(wb, sku_df):
    ws = wb.create_sheet("Grooming SKU Detail")
    _style_title(ws, 1, "Grooming SKU Detail", merge_to=8)
    _style_subtitle(ws, 2, "All grooming service SKUs ranked by revenue", merge_to=8)

    row = 4
    headers = ["Rank", "SKU", "Service Name", "Service Type", "Units", "Net Sales", "% of Service Mix", "Avg Ticket"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 8)

    for rank, (_, r) in enumerate(sku_df.head(50).iterrows(), 1):
        row += 1
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=str(r["sku"]))
        ws.cell(row=row, column=3, value=r["name"])
        ws.cell(row=row, column=4, value=r["service_type"])
        ws.cell(row=row, column=5, value=_fmt_int(r["units"]))
        ws.cell(row=row, column=6, value=_fmt_currency(r["net_sales"]))
        ws.cell(row=row, column=7, value=_fmt_pct(r["pct_mix"]))
        ws.cell(row=row, column=8, value=_fmt_currency(r["avg_ticket"]))

    _zebra_rows(ws, 5, row, 8)
    _auto_width(ws)
    return ws


def write_top_retail(wb, top_rev, top_units):
    for sheet_name, data, sort_col in [
        ("Top 50 Retail (by $)", top_rev, "net_sales"),
        ("Top 50 Retail (by Units)", top_units, "units"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _style_title(ws, 1, sheet_name, merge_to=8)
        _style_subtitle(ws, 2, f"Top 50 retail SKUs ranked by {'net sales' if '$' in sheet_name else 'units sold'}", merge_to=8)

        row = 4
        headers = ["Rank", "SKU", "Product Name", "Category", "Units", "Net Sales", "% of Retail", "Avg Price"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, 8)

        for rank, (_, r) in enumerate(data.iterrows(), 1):
            row += 1
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=str(r["sku"]))
            ws.cell(row=row, column=3, value=r["name"])
            ws.cell(row=row, column=4, value=r["retail_category"])
            ws.cell(row=row, column=5, value=_fmt_int(r["units"]))
            ws.cell(row=row, column=6, value=_fmt_currency(r["net_sales"]))
            ws.cell(row=row, column=7, value=_fmt_pct(r["pct_retail"]))
            ws.cell(row=row, column=8, value=_fmt_currency(r["avg_price"]))

        _zebra_rows(ws, 5, row, 8)
        _auto_width(ws)


def write_monthly_performance(wb, monthly_df):
    ws = wb.create_sheet("Monthly Performance")
    _style_title(ws, 1, "Monthly Performance Dashboard", merge_to=9)
    _style_subtitle(ws, 2, f"{STORE_NAME} -- {START_DATE} to {END_DATE}", merge_to=9)

    row = 4
    ws.cell(row=row, column=1, value="Total Store Revenue by Month").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    headers = ["Month", "Net Revenue", "Grooming Rev", "Retail Rev", "Groom %", "Retail %", "Tickets", "Avg Ticket", "MoM Growth"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 9)

    for _, r in monthly_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["month_name"])
        ws.cell(row=row, column=2, value=_fmt_currency(r["net_revenue"]))
        ws.cell(row=row, column=3, value=_fmt_currency(r["grooming_rev"]))
        ws.cell(row=row, column=4, value=_fmt_currency(r["retail_rev"]))
        ws.cell(row=row, column=5, value=_fmt_pct(r["groom_pct"]))
        ws.cell(row=row, column=6, value=_fmt_pct(r["retail_pct"]))
        ws.cell(row=row, column=7, value=int(r["tickets"]))
        ws.cell(row=row, column=8, value=_fmt_currency(r["avg_ticket"]))
        ws.cell(row=row, column=9, value=_fmt_pct(r["mom_growth"]) if not pd.isna(r["mom_growth"]) else "--")

    _zebra_rows(ws, 7, row, 9)

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=_fmt_currency(monthly_df["net_revenue"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=3, value=_fmt_currency(monthly_df["grooming_rev"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=4, value=_fmt_currency(monthly_df["retail_rev"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=7, value=int(monthly_df["tickets"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=8, value=_fmt_currency(monthly_df["net_revenue"].sum()/monthly_df["tickets"].sum())).font = Font(bold=True)

    if len(monthly_df) >= 2:
        jan = monthly_df.iloc[0]["net_revenue"]
        dec = monthly_df.iloc[-1]["net_revenue"]
        if jan > 0:
            growth = (dec - jan) / jan * 100
            row += 2
            ws.cell(row=row, column=1, value=f"Full year growth: {growth:.1f}% from {monthly_df.iloc[0]['month_name']} to {monthly_df.iloc[-1]['month_name']}. Avg ticket grew from {_fmt_currency(monthly_df.iloc[0]['avg_ticket'])} to {_fmt_currency(monthly_df.iloc[-1]['avg_ticket'])}.").font = Font(name="Inter", size=10, italic=True)

    _auto_width(ws)

    try:
        chart = LineChart()
        chart.title = "Monthly Revenue Trend"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Month"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        data_start_row = row + 3
        for i, (_, r) in enumerate(monthly_df.iterrows()):
            ws.cell(row=data_start_row + i, column=11, value=r["month_name"])
            ws.cell(row=data_start_row + i, column=12, value=r["net_revenue"])
            ws.cell(row=data_start_row + i, column=13, value=r["grooming_rev"])
            ws.cell(row=data_start_row + i, column=14, value=r["retail_rev"])

        cats = Reference(ws, min_col=11, min_row=data_start_row, max_row=data_start_row + 11)
        data_ref = Reference(ws, min_col=12, max_col=14, min_row=data_start_row - 1, max_row=data_start_row + 11)

        ws.cell(row=data_start_row - 1, column=12, value="Total")
        ws.cell(row=data_start_row - 1, column=13, value="Grooming")
        ws.cell(row=data_start_row - 1, column=14, value="Retail")

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row + 3}")
    except Exception as e:
        print(f"  Chart error: {e}")

    return ws


def write_top_customers(wb, top_cust):
    ws = wb.create_sheet("Top Customers")
    _style_title(ws, 1, "Top Customers -- Ranked by Total Spend", merge_to=7)
    _style_subtitle(ws, 2, "Top 50 customers by total spend", merge_to=7)

    row = 4
    headers = ["Customer ID", "Total Spend", "Transactions", "Visit Days", "Spend/Visit", "Avg Txn", "Tier"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 7)

    tier_colors = {
        "Premium": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "High": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "Standard": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        "Value": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    }

    for _, r in top_cust.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=int(r["customer_id"]))
        ws.cell(row=row, column=2, value=_fmt_currency(r["total_spend"]))
        ws.cell(row=row, column=3, value=int(r["transactions"]))
        ws.cell(row=row, column=4, value=int(r["visit_days"]))
        ws.cell(row=row, column=5, value=_fmt_currency(r["spend_per_visit"]))
        ws.cell(row=row, column=6, value=_fmt_currency(r["avg_txn"]))
        ws.cell(row=row, column=7, value=r["tier"])
        if r["tier"] in tier_colors:
            ws.cell(row=row, column=7).fill = tier_colors[r["tier"]]

    _auto_width(ws)
    return ws


def write_customer_intelligence(wb, ci):
    ws = wb.create_sheet("Customer Intelligence")
    _style_title(ws, 1, "Customer Intelligence Dashboard", merge_to=8)
    _style_subtitle(ws, 2, f"{STORE_NAME} -- {START_DATE} to {END_DATE}", merge_to=8)

    row = 4
    ws.cell(row=row, column=1, value="Ticket Composition Analysis").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    for i, h in enumerate(["Ticket Type", "Tickets", "% of Tickets", "Revenue", "% of Revenue", "Avg Ticket"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 6)

    for _, r in ci["ticket_composition"].iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["ticket_type"])
        ws.cell(row=row, column=2, value=_fmt_int(r["tickets"]))
        ws.cell(row=row, column=3, value=_fmt_pct(r["pct_tickets"]))
        ws.cell(row=row, column=4, value=_fmt_currency(r["revenue"]))
        ws.cell(row=row, column=5, value=_fmt_pct(r["pct_revenue"]))
        ws.cell(row=row, column=6, value=_fmt_currency(r["avg_ticket"]))

    row += 3
    ws.cell(row=row, column=1, value="Customer Visit Frequency").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    for i, h in enumerate(["Visit Frequency", "Customers", "% of Customers", "Revenue", "% of Revenue", "Avg Spend/Customer"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 6)

    freq_order = ["1 visit (trial)", "2 visits", "3-4 visits (developing)", "5-8 visits (loyal)", "9-12 visits (committed)", "13+ visits (champions)"]
    freq_df = ci["visit_frequency"]
    freq_df["_order"] = freq_df["bucket"].map({s: i for i, s in enumerate(freq_order)})
    freq_df = freq_df.sort_values("_order")

    for _, r in freq_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["bucket"])
        ws.cell(row=row, column=2, value=_fmt_int(r["customers"]))
        ws.cell(row=row, column=3, value=_fmt_pct(r["pct_customers"]))
        ws.cell(row=row, column=4, value=_fmt_currency(r["revenue"]))
        ws.cell(row=row, column=5, value=_fmt_pct(r["pct_revenue"]))
        ws.cell(row=row, column=6, value=_fmt_currency(r["avg_spend"]))

    row += 3
    ws.cell(row=row, column=1, value="Customer Recency & Churn Risk").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    _style_subtitle(ws, row + 1, "Days since last visit as of Dec 31, 2025", merge_to=5)
    row += 2
    for i, h in enumerate(["Segment", "Customers", "%", "Total Spend", "Avg Spend"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 5)

    seg_order = ["Active (last 30 days)", "Recent (31-60 days)", "At Risk (61-90 days)", "Lapsing (91-180 days)", "Lost (180+ days)"]
    churn_df = ci["churn_risk"]
    churn_df["_order"] = churn_df["segment"].map({s: i for i, s in enumerate(seg_order)})
    churn_df = churn_df.sort_values("_order")

    for _, r in churn_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["segment"])
        ws.cell(row=row, column=2, value=_fmt_int(r["customers"]))
        ws.cell(row=row, column=3, value=_fmt_pct(r["pct"]))
        ws.cell(row=row, column=4, value=_fmt_currency(r["total_spend"]))
        ws.cell(row=row, column=5, value=_fmt_currency(r["avg_spend"]))

    row += 3
    ws.cell(row=row, column=1, value="Day of Week Performance").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    for i, h in enumerate(["Day", "Revenue", "% of Week", "Tickets", "Avg Ticket", "Grooming Rev", "Retail Rev", "Retail %"], 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, 8)

    for _, r in ci["day_of_week"].iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["day"])
        ws.cell(row=row, column=2, value=_fmt_currency(r["revenue"]))
        ws.cell(row=row, column=3, value=_fmt_pct(r["pct_week"]))
        ws.cell(row=row, column=4, value=int(r["tickets"]))
        ws.cell(row=row, column=5, value=_fmt_currency(r["avg_ticket"]))
        ws.cell(row=row, column=6, value=_fmt_currency(r["grooming_rev"]))
        ws.cell(row=row, column=7, value=_fmt_currency(r["retail_rev"]))
        ws.cell(row=row, column=8, value=_fmt_pct(r["retail_pct"]))

    _auto_width(ws)
    return ws


def write_team_performance(wb, team_df):
    ws = wb.create_sheet("Team Performance")
    _style_title(ws, 1, "Team Performance", merge_to=12)
    _style_subtitle(ws, 2, "Groomer scorecards and productivity metrics", merge_to=12)

    row = 4
    ws.cell(row=row, column=1, value="Groomer Scorecard").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    headers = ["Groomer", "Revenue", "Appointments", "$/Appt", "Core Visits",
               "SPA %", "Add-On %", "Teeth %", "Hours", "$/Hour"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))

    for _, r in team_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["salesperson"])
        ws.cell(row=row, column=2, value=_fmt_currency(r["revenue"]))
        ws.cell(row=row, column=3, value=int(r["appointments"]))
        ws.cell(row=row, column=4, value=_fmt_currency(r["per_appt"]))
        ws.cell(row=row, column=5, value=_fmt_int(r["core_visits"]))
        ws.cell(row=row, column=6, value=_fmt_pct(r["spa_pct"]))
        ws.cell(row=row, column=7, value=_fmt_pct(r["addon_pct"]))
        ws.cell(row=row, column=8, value=_fmt_pct(r["teeth_pct"]))
        ws.cell(row=row, column=9, value=f"{r['hours']:.1f}" if not pd.isna(r.get("hours")) else "N/A")
        ws.cell(row=row, column=10, value=_fmt_currency(r["per_hour"]) if not pd.isna(r.get("per_hour")) else "N/A")

        spa_val = r["spa_pct"]
        if spa_val >= 40:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif spa_val >= 25:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        else:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    _zebra_rows(ws, 7, row, len(headers))

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=_fmt_currency(team_df["revenue"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=3, value=int(team_df["appointments"].sum())).font = Font(bold=True)

    _auto_width(ws)
    return ws


def write_category_top10s(wb, cat_data):
    ws = wb.create_sheet("Category Top 10s")
    _style_title(ws, 1, "Category Top 10s", merge_to=10)
    _style_subtitle(ws, 2, "Top 10 by revenue (left) and by units (right) for each retail category", merge_to=10)

    row = 4
    for cat_name in sorted(cat_data.keys(), key=lambda x: cat_data[x]["total"], reverse=True):
        info = cat_data[cat_name]
        ws.cell(row=row, column=1, value=f"{cat_name}  |  Total Revenue: {_fmt_currency(info['total'])}").font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        row += 1

        for i, h in enumerate(["Rank", "Product Name", "Revenue", "Units"], 1):
            ws.cell(row=row, column=i, value=h)
        for i, h in enumerate(["Rank", "Product Name", "Units", "Revenue"], 1):
            ws.cell(row=row, column=i + 5, value=h)
        _style_header(ws, row, 4)
        for c in range(6, 10):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
            ws.cell(row=row, column=c).font = Font(name="Inter", size=10, bold=True, color=WHITE)

        top_rev = info["top_by_revenue"]
        top_units = info["top_by_units"]

        max_rows = max(len(top_rev), len(top_units))
        for i in range(max_rows):
            row += 1
            if i < len(top_rev):
                r = top_rev.iloc[i]
                ws.cell(row=row, column=1, value=i + 1)
                ws.cell(row=row, column=2, value=r["name"])
                ws.cell(row=row, column=3, value=_fmt_currency(r["revenue"]))
                ws.cell(row=row, column=4, value=_fmt_int(r["units"]))
            if i < len(top_units):
                r = top_units.iloc[i]
                ws.cell(row=row, column=6, value=i + 1)
                ws.cell(row=row, column=7, value=r["name"])
                ws.cell(row=row, column=8, value=_fmt_int(r["units"]))
                ws.cell(row=row, column=9, value=_fmt_currency(r["revenue"]))

        row += 3

    _auto_width(ws)
    return ws


def write_brand_breakdown(wb, brand_data):
    ws = wb.create_sheet("Brand Breakdown")
    _style_title(ws, 1, "Brand Breakdown by Category", merge_to=7)
    _style_subtitle(ws, 2, "Top 10 brands per retail category, sorted by revenue", merge_to=7)

    row = 4
    for cat_name in sorted(brand_data.keys(), key=lambda x: brand_data[x]["total"], reverse=True):
        info = brand_data[cat_name]
        ws.cell(row=row, column=1, value=f"{cat_name}  |  Total: {_fmt_currency(info['total'])}").font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        row += 1

        headers = ["Brand", "Net Sales", "% of Category", "Units", "SKU Count", "Avg Price/Item", "Revenue/SKU"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        _style_header(ws, row, 7)

        for _, r in info["brands"].iterrows():
            row += 1
            ws.cell(row=row, column=1, value=r["brand"])
            ws.cell(row=row, column=2, value=_fmt_currency(r["net_sales"]))
            ws.cell(row=row, column=3, value=_fmt_pct(r["pct_category"]))
            ws.cell(row=row, column=4, value=_fmt_int(r["units"]))
            ws.cell(row=row, column=5, value=int(r["sku_count"]))
            ws.cell(row=row, column=6, value=_fmt_currency(r["avg_price"]))
            ws.cell(row=row, column=7, value=_fmt_currency(r["revenue_per_sku"]))

        row += 3

    _auto_width(ws)
    return ws


def generate_workbook(transformed):
    print("\n" + "=" * 60)
    print("GENERATING EXCEL WORKBOOK")
    print("=" * 60)

    df = transformed["df_items"]
    df_orders = transformed["df_orders"]
    df_clocks = transformed["df_clocks"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  [1/10] Executive Summary...")
    summary = compute_executive_summary(df, df_orders)
    write_executive_summary(wb, summary, df)

    print("  [2/10] Service Type Mix...")
    mix = compute_service_mix(df)
    if not mix.empty:
        write_service_mix(wb, mix)

    print("  [3/10] Dog Size Breakdown...")
    all_sizes, std_sizes, dood_sizes = compute_dog_sizes(df)
    if isinstance(all_sizes, pd.DataFrame) and not all_sizes.empty:
        write_dog_sizes(wb, all_sizes, std_sizes, dood_sizes)

    print("  [4/10] Grooming SKU Detail...")
    groom_skus = compute_grooming_skus(df)
    if not groom_skus.empty:
        write_grooming_skus(wb, groom_skus)

    print("  [5/10] Top 50 Retail...")
    top_rev, top_units = compute_top_retail(df)
    if not top_rev.empty:
        write_top_retail(wb, top_rev, top_units)

    print("  [6/10] Category Top 10s...")
    cat_data = compute_category_top10s(df)
    if cat_data:
        write_category_top10s(wb, cat_data)

    print("  [7/10] Brand Breakdown...")
    brand_data = compute_brand_breakdown(df)
    if brand_data:
        write_brand_breakdown(wb, brand_data)

    print("  [8/10] Monthly Performance...")
    monthly = compute_monthly_performance(df, df_orders)
    write_monthly_performance(wb, monthly)

    print("  [9/10] Top Customers...")
    top_cust = compute_top_customers(df, df_orders)
    write_top_customers(wb, top_cust)

    print("  [10/10] Customer Intelligence...")
    ci = compute_customer_intelligence(df, df_orders)
    write_customer_intelligence(wb, ci)

    print("  [11/11] Team Performance...")
    team = compute_team_performance(df, df_orders, df_clocks)
    write_team_performance(wb, team)

    _store_short = STORE_NAME.split("--")[-1].strip().split("(")[0].strip().replace(", ", "_").replace(" ", "")
    _year_range = f"{START_DATE[:4]}" if START_DATE[:4] == END_DATE[:4] else f"{START_DATE[:4]}-{END_DATE[:4]}"
    desktop = Path.home() / "Desktop"
    desktop.mkdir(exist_ok=True)
    output_file = desktop / f"WoofGang_{_store_short}_{_year_range}_Analysis.xlsx"
    wb.save(output_file)
    print(f"\n{'=' * 60}")
    print(f"REPORT SAVED: {output_file}")
    print(f"{'=' * 60}")

    print(f"\nStore: {STORE_NAME}")
    print(f"Period: {START_DATE} to {END_DATE}")
    print(f"Total Net Sales: {_fmt_currency(summary['total_net_sales'])}")
    print(f"  Grooming: {_fmt_currency(summary['grooming_revenue'])} ({summary['grooming_pct']:.1f}%)")
    print(f"  Retail: {_fmt_currency(summary['retail_revenue'])} ({summary['retail_pct']:.1f}%)")
    print(f"Total Transactions: {_fmt_int(summary['total_transactions'])}")
    print(f"Avg Transaction: {_fmt_currency(summary['avg_transaction_value'])}")
    print(f"Sheets: {len(wb.sheetnames)}")

    return output_file


# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    start_time = time.time()
    raw_data = extract_all_data()
    transformed = transform_data(raw_data)
    output = generate_workbook(transformed)
    elapsed = time.time() - start_time
    print(f"\nTotal time: {elapsed:.0f}s")



def extract_full_catalog():
    """Pull full product catalog from FranPOS including stock levels."""
    print("Fetching full product catalog from FranPOS...")
    all_products = []
    seen = set()
    page = 1
    while True:
        try:
            data = api_get("getProductServicesByCompany", {
                "locationId": LOCATION_ID,
                "page": page
            })
        except Exception as e:
            print(f"  Catalog page {page} error: {e}")
            break
        if not data:
            break
        items = data if isinstance(data, list) else data.get("ProductServices", data.get("Items", data.get("Data", [])))
        if not items:
            break
        new = 0
        for item in items:
            uid = str(item.get("Id","")) + str(item.get("Sku",""))
            if uid not in seen:
                seen.add(uid)
                all_products.append(item)
                new += 1
        print(f"  Page {page}: {new} products (total: {len(all_products)})")
        if new == 0 or len(items) < 50:
            break
        page += 1
    import json as _json
    out = DATA_DIR / "full_catalog.json"
    with open(out, "w") as f:
        _json.dump(all_products, f, indent=2)
    print(f"  Catalog saved: {len(all_products)} products")
    return all_products


def generate_dashboard():
    """Generate the tabbed HTML dashboard for all years."""
    import sys, types, json as _json, pandas as _pd
    from pathlib import Path as _Path

    print("\nGenerating HTML dashboard...")

    _fake = types.ModuleType("run")
    _fake.STORE_NAME = STORE_NAME
    _fake.START_DATE = START_DATE
    _fake.END_DATE   = "2026-12-31"
    _fake.DATA_DIR   = DATA_DIR
    _fake.classify_item = classify_item
    sys.modules["run"] = _fake

    _scripts_dir = Path(__file__).parent
    sys.path.insert(0, str(_scripts_dir))
    import generate_dashboards as gd

    # all_data.json is already written by extract_all_data() above
    pass

    df_all, df_orders_all, _ = gd.load_data()

    YEARS = {
        "2024": ("2024-09-26", "2024-12-31"),
        "2025": ("2025-01-01", "2025-12-31"),
        "2026": ("2026-01-01", "2026-12-31"),
    }
    YEAR_LABELS = {
        "2024": "2024 (Sep\u2013Dec)",
        "2025": "2025 (Full Year)",
        "2026": "2026 (YTD)",
    }

    bodies = {}
    for yr, (start, end) in YEARS.items():
        _fake.START_DATE = start
        _fake.END_DATE = end
        df_yr = df_all[df_all["year"] == int(yr)]
        df_ord_yr = df_orders_all[df_orders_all["year"] == int(yr)]
        bodies[yr] = gd.generate_main_dashboard(df_yr, df_ord_yr, None, body_only=True, year_suffix=yr)

    _last_txn_all = df_all["created"].max()
    _through_all = _last_txn_all.strftime("%-m/%-d/%Y") if not _pd.isnull(_last_txn_all) else "today"

    tabbed_css = """
.yr-tab-bar { background: #C4276E; padding: 0 24px; display: flex; gap: 4px; position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 8px rgba(0,0,0,0.25); }
.yr-tab { padding: 14px 32px; border: none; background: transparent; color: rgba(255,255,255,0.6); font-size: 15px; font-weight: 600; cursor: pointer; border-bottom: 3px solid transparent; transition: all 0.2s; font-family: inherit; }
.yr-tab:hover { color: white; background: rgba(255,255,255,0.08); }
.yr-tab.active { color: white; border-bottom-color: white; }
"""
    tabbed_js = """
var _chartsInited = {};
function showYear(yr) {
    document.querySelectorAll('.yr-panel').forEach(function(p) { p.style.display = 'none'; });
    document.querySelectorAll('.yr-tab').forEach(function(t) { t.classList.remove('active'); });
    document.getElementById('panel-' + yr).style.display = 'block';
    document.getElementById('tab-' + yr).classList.add('active');
    if (!_chartsInited[yr]) {
        _chartsInited[yr] = true;
        if (yr === '2024' && typeof initCharts_2024 === 'function') initCharts_2024();
        if (yr === '2025' && typeof initCharts_2025 === 'function') initCharts_2025();
        if (yr === '2026' && typeof initCharts_2026 === 'function') initCharts_2026();
    }
}
window.addEventListener('DOMContentLoaded', function() { showYear('2025'); });
"""

    import re as _re
    html = gd.html_head("Woof Gang Port Washington", f"Store Performance Analysis \u00b7 Sales through {_through_all}")
    html = html.replace("</style>", tabbed_css + "\n</style>", 1)
    html += '<div class="yr-tab-bar">\n'
    for yr in ["2024", "2025", "2026"]:
        active = "active" if yr == "2025" else ""
        html += f'  <button class="yr-tab {active}" onclick="showYear(\'{yr}\')" id="tab-{yr}">{YEAR_LABELS[yr]}</button>\n'
    html += '</div>\n'

    for yr in ["2024", "2025", "2026"]:
        display = "block" if yr == "2025" else "none"
        html += f'<div class="yr-panel" id="panel-{yr}" style="display:{display}">\n'
        body = bodies[yr]
        def _wrap(m, _yr=yr):
            inner = m.group(0)[len("<script>\n"):-len("</script>")]
            return f"<script>\nfunction initCharts_{_yr}() {{\n{inner}\n}}\n</script>"
        body = _re.sub(r'<script>\nconst cc = .*?</script>', _wrap, body, flags=_re.DOTALL)
        html += body
        html += '</div>\n'

    html += f'<script>\n{tabbed_js}\n</script>\n'
    html += gd.HTML_FOOT

    out_path = OUTPUT_DIR / "WoofGang_PortWashington_NY_AllYears_Dashboard.html"
    with open(out_path, "w") as f:
        f.write(html)
    print(f"  Dashboard saved: {out_path}")
    return out_path

try:
    extract_full_catalog()
except Exception as e:
    print(f"Catalog error: {e}")
try:
    generate_dashboard()
except Exception as e:
    import traceback
    print(f'Dashboard error: {e}')
    traceback.print_exc()


def reset_untracked_stock():
    """Reset bulk/treat SKUs that can't be inventory-tracked back to 0 nightly."""
    import httpx as _httpx, time as _time
    UNTRACKED = [
        ("TREAT02", "Farmers Market Bulk"),
        ("TREAT01", "Bakery Bay Bulk Treats"),
        ("1002", "Cookie Large"),
        ("1003", "Cookie XL"),
        ("1004", "Cookie Small"),
        ("1005", "Cookie Medium"),
        ("1006", "Cookie XS"),
        ("2001", "Farmers Market Bulk Treats"),
        ("2002", "Bakery Bulk Treats"),
        ("82101", "Preppy Puppy Birthday Cupcakes"),
        ("81909", "Preppy Puppy Blue 4\" Birthday Cake"),
        ("000", "Miscellaneous"),
    ]
    print("\nResetting untracked SKUs to 0...")
    for sku, name in UNTRACKED:
        r = _httpx.post(f"{BASE_URL}/api/updateStockByProductSKU",
                        params={"sku": sku, "stock": 0, "addToStock": "false", "Token": TOKEN},
                        timeout=15)
        ok = r.status_code == 200 and r.json()[0].get("IsSuccess")
        print(f"  {'✓' if ok else '✗'} {sku:12} {name}")
        _time.sleep(0.3)
    print("  Done!")

try:
    reset_untracked_stock()
except Exception as e:
    print(f"Stock reset error: {e}")

try:
    import subprocess
    result = subprocess.run(["python3", str(Path(__file__).parent / "build_commission_dashboard.py")],
                           capture_output=True, text=True)
    print(result.stdout)
    if result.returncode != 0:
        print(f"Commission dashboard error: {result.stderr}")
except Exception as e:
    print(f"Commission dashboard error: {e}")
