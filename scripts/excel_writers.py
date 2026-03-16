"""
Excel worksheet writers: each function creates one sheet in the workbook.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference

from config import PAW_MAGENTA, TEDDY_BROWN, DARK_TEAL, WHITE
from formatting import fmt_currency, fmt_pct, fmt_int
from excel_styles import style_header, style_title, style_subtitle, auto_width, zebra_rows


def write_executive_summary(wb, summary, df_items, *, store_name, start_date, end_date):
    ws = wb.create_sheet("Executive Summary")
    style_title(ws, 1, store_name, merge_to=5)
    style_subtitle(ws, 2, f"Performance Analysis  |  {start_date} to {end_date}", merge_to=5)

    row = 4
    ws.cell(row=row, column=1, value="Store Overview").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 1
    for i, h in enumerate(["Metric", "Value"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 2)

    metrics = [
        ("Total Net Sales", fmt_currency(summary["total_net_sales"])),
        ("Total Gross Sales", fmt_currency(summary["total_gross_sales"])),
        ("Total Returns", fmt_currency(summary["total_returns"])),
        ("Total Discounts", fmt_currency(summary["total_discounts"])),
        ("Total Units Sold", fmt_int(summary["total_units"])),
        ("Total Transactions", fmt_int(summary["total_transactions"])),
        ("Total SKUs", fmt_int(summary["total_skus"])),
        ("Avg Price per Item", fmt_currency(summary["avg_price_per_item"])),
        ("Average Transaction Value", fmt_currency(summary["avg_transaction_value"])),
        ("Date Range", f"{start_date} to {end_date}"),
    ]
    for i, (metric, val) in enumerate(metrics):
        ws.cell(row=row + 1 + i, column=1, value=metric).font = Font(name="Inter", size=10)
        ws.cell(row=row + 1 + i, column=2, value=val).font = Font(name="Inter", size=10, bold=True)
    zebra_rows(ws, row + 1, row + len(metrics), 2)

    row += len(metrics) + 2
    ws.cell(row=row, column=1, value="Revenue Mix").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 1
    for i, h in enumerate(["Category", "Revenue", "% of Total"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 3)

    mix_data = [
        ("Grooming Services", summary["grooming_revenue"], summary["grooming_pct"]),
        ("Retail Sales", summary["retail_revenue"], summary["retail_pct"]),
        ("Gift Cards", summary["gift_card_revenue"],
         summary["gift_card_revenue"] / summary["total_net_sales"] * 100 if summary["total_net_sales"] else 0),
    ]
    for i, (cat, rev, pct) in enumerate(mix_data):
        ws.cell(row=row + 1 + i, column=1, value=cat)
        ws.cell(row=row + 1 + i, column=2, value=fmt_currency(rev))
        ws.cell(row=row + 1 + i, column=3, value=fmt_pct(pct))
    zebra_rows(ws, row + 1, row + len(mix_data), 3)

    groom = df_items[df_items["is_groom"] == True]
    if not groom.empty:
        row += len(mix_data) + 2
        ws.cell(row=row, column=1, value="SKU Concentration -- Grooming Services").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
        style_subtitle(ws, row + 1, "Pareto/ABC analysis of grooming service SKUs only", merge_to=5)
        row += 2
        for i, h in enumerate(["Class", "SKU Count", "% of SKUs", "Revenue", "% of Revenue"], 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, 5)

        sku_rev = groom.groupby("sku")["net_sales"].sum().sort_values(ascending=False)
        total_groom_rev = sku_rev.sum()
        cumsum = sku_rev.cumsum() / total_groom_rev
        a_count = (cumsum <= 0.80).sum()
        b_count = ((cumsum > 0.80) & (cumsum <= 0.95)).sum()
        c_count = (cumsum > 0.95).sum()
        total_skus = len(sku_rev)

        abc = [
            ("A-Class (Top 80%)", a_count, a_count/total_skus*100, sku_rev.iloc[:a_count].sum(), sku_rev.iloc[:a_count].sum()/total_groom_rev*100),
            ("B-Class (Next 15%)", b_count, b_count/total_skus*100, sku_rev.iloc[a_count:a_count+b_count].sum(), sku_rev.iloc[a_count:a_count+b_count].sum()/total_groom_rev*100),
            ("C-Class (Remaining 5%)", c_count, c_count/total_skus*100, sku_rev.iloc[a_count+b_count:].sum(), sku_rev.iloc[a_count+b_count:].sum()/total_groom_rev*100),
        ]
        for i, (cls, cnt, pct_s, rev, pct_r) in enumerate(abc):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=cls)
            ws.cell(row=r, column=2, value=cnt)
            ws.cell(row=r, column=3, value=fmt_pct(pct_s))
            ws.cell(row=r, column=4, value=fmt_currency(rev))
            ws.cell(row=r, column=5, value=fmt_pct(pct_r))
        zebra_rows(ws, row + 1, row + 3, 5)

    retail = df_items[df_items["is_retail"] == True]
    if not retail.empty:
        row += 5
        ws.cell(row=row, column=1, value="SKU Concentration -- Retail").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
        style_subtitle(ws, row + 1, "Pareto/ABC analysis of retail product SKUs only", merge_to=5)
        row += 2
        for i, h in enumerate(["Class", "SKU Count", "% of SKUs", "Revenue", "% of Revenue"], 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, 5)

        sku_rev = retail.groupby("sku")["net_sales"].sum().sort_values(ascending=False)
        total_retail_rev = sku_rev.sum()
        cumsum = sku_rev.cumsum() / total_retail_rev
        a_count = max((cumsum <= 0.80).sum(), 1)
        b_count = max(((cumsum > 0.80) & (cumsum <= 0.95)).sum(), 0)
        c_count = max((cumsum > 0.95).sum(), 0)
        total_skus = len(sku_rev)

        abc = [
            ("A-Class (Top 80%)", a_count, a_count/total_skus*100 if total_skus else 0,
             sku_rev.iloc[:a_count].sum(), sku_rev.iloc[:a_count].sum()/total_retail_rev*100 if total_retail_rev else 0),
            ("B-Class (Next 15%)", b_count, b_count/total_skus*100 if total_skus else 0,
             sku_rev.iloc[a_count:a_count+b_count].sum(),
             sku_rev.iloc[a_count:a_count+b_count].sum()/total_retail_rev*100 if total_retail_rev else 0),
            ("C-Class (Remaining 5%)", c_count, c_count/total_skus*100 if total_skus else 0,
             sku_rev.iloc[a_count+b_count:].sum(),
             sku_rev.iloc[a_count+b_count:].sum()/total_retail_rev*100 if total_retail_rev else 0),
        ]
        for i, (cls, cnt, pct_s, rev, pct_r) in enumerate(abc):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=cls)
            ws.cell(row=r, column=2, value=cnt)
            ws.cell(row=r, column=3, value=fmt_pct(pct_s))
            ws.cell(row=r, column=4, value=fmt_currency(rev))
            ws.cell(row=r, column=5, value=fmt_pct(pct_r))
        zebra_rows(ws, row + 1, row + 3, 5)

    auto_width(ws)
    return ws


def write_service_mix(wb, mix_df):
    ws = wb.create_sheet("Service Type Mix")
    style_title(ws, 1, "Service Type Mix", merge_to=7)
    style_subtitle(ws, 2, "Grooming service breakdown using actual location service names", merge_to=7)

    row = 4
    headers = ["Service Type", "Units", "% of Units", "Revenue", "% of Revenue", "Avg Ticket", "SKU Count"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 7)

    for i, (_, r) in enumerate(mix_df.iterrows()):
        row += 1
        ws.cell(row=row, column=1, value=r["service_type"])
        ws.cell(row=row, column=2, value=fmt_int(r["units"]))
        ws.cell(row=row, column=3, value=fmt_pct(r["pct_units"]))
        ws.cell(row=row, column=4, value=fmt_currency(r["revenue"]))
        ws.cell(row=row, column=5, value=fmt_pct(r["pct_revenue"]))
        ws.cell(row=row, column=6, value=fmt_currency(r["avg_ticket"]))
        ws.cell(row=row, column=7, value=int(r["sku_count"]))

    zebra_rows(ws, 5, row, 7)

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=fmt_int(mix_df["units"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=3, value="100.0%").font = Font(bold=True)
    ws.cell(row=row, column=4, value=fmt_currency(mix_df["revenue"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=5, value="100.0%").font = Font(bold=True)
    ws.cell(row=row, column=6, value=fmt_currency(mix_df["revenue"].sum() / mix_df["units"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=7, value=int(mix_df["sku_count"].sum())).font = Font(bold=True)

    auto_width(ws)
    return ws


def write_dog_sizes(wb, all_sizes, std_sizes, dood_sizes):
    ws = wb.create_sheet("Dog Size Breakdown")
    style_title(ws, 1, "All Dogs -- Size Distribution", merge_to=6)
    style_subtitle(ws, 2, "Core grooms (Full Groom + Mini Groom + Luxury Bath + Classic Bath)", merge_to=6)

    def write_size_table(ws, start_row, title, data):
        ws.cell(row=start_row, column=1, value=title).font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        r = start_row + 1
        for i, h in enumerate(["Size Bracket", "Units", "% of Units", "Revenue", "% of Revenue", "Avg Ticket"], 1):
            ws.cell(row=r, column=i, value=h)
        style_header(ws, r, 6)

        for idx, row_data in data.iterrows():
            r += 1
            ws.cell(row=r, column=1, value=row_data["dog_size"])
            ws.cell(row=r, column=2, value=fmt_int(row_data["units"]))
            ws.cell(row=r, column=3, value=fmt_pct(row_data.get("pct_units", row_data["units"]/data["units"].sum()*100)))
            ws.cell(row=r, column=4, value=fmt_currency(row_data["revenue"]))
            ws.cell(row=r, column=5, value=fmt_pct(row_data.get("pct_revenue", row_data["revenue"]/data["revenue"].sum()*100)))
            ws.cell(row=r, column=6, value=fmt_currency(row_data["avg_ticket"]))

        zebra_rows(ws, start_row + 2, r, 6)
        r += 1
        ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=r, column=2, value=fmt_int(data["units"].sum())).font = Font(bold=True)
        ws.cell(row=r, column=4, value=fmt_currency(data["revenue"].sum())).font = Font(bold=True)
        ws.cell(row=r, column=6, value=fmt_currency(data["revenue"].sum()/data["units"].sum() if data["units"].sum() else 0)).font = Font(bold=True)
        return r + 2

    next_row = 4
    if isinstance(all_sizes, pd.DataFrame) and not all_sizes.empty:
        next_row = write_size_table(ws, next_row, "All Dogs", all_sizes)
    if isinstance(std_sizes, pd.DataFrame) and not std_sizes.empty:
        next_row = write_size_table(ws, next_row, "Standard Dogs (Non-Doodle/Poodle)", std_sizes)
    if isinstance(dood_sizes, pd.DataFrame) and not dood_sizes.empty:
        next_row = write_size_table(ws, next_row, "Doodle Dogs (Poodle & Doodle breeds)", dood_sizes)

    if isinstance(std_sizes, pd.DataFrame) and not std_sizes.empty and isinstance(dood_sizes, pd.DataFrame) and not dood_sizes.empty:
        ws.cell(row=next_row, column=1, value="Standard vs Doodle Price Comparison").font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        r = next_row + 1
        for i, h in enumerate(["Size", "Std Avg Ticket", "Doodle Avg Ticket", "Premium $", "Premium %"], 1):
            ws.cell(row=r, column=i, value=h)
        style_header(ws, r, 5)

        for size in ["0-20 lbs", "21-40 lbs", "41-75 lbs", "76-100 lbs", "Over 100 lbs"]:
            std_row = std_sizes[std_sizes["dog_size"] == size]
            dood_row = dood_sizes[dood_sizes["dog_size"] == size]
            if std_row.empty or dood_row.empty:
                continue
            std_avg = std_row["avg_ticket"].values[0]
            dood_avg = dood_row["avg_ticket"].values[0]
            premium = dood_avg - std_avg
            prem_pct = premium / std_avg * 100 if std_avg else 0
            r += 1
            ws.cell(row=r, column=1, value=size)
            ws.cell(row=r, column=2, value=fmt_currency(std_avg))
            ws.cell(row=r, column=3, value=fmt_currency(dood_avg))
            ws.cell(row=r, column=4, value=fmt_currency(premium))
            ws.cell(row=r, column=5, value=fmt_pct(prem_pct))

    auto_width(ws)
    return ws


def write_grooming_skus(wb, sku_df):
    ws = wb.create_sheet("Grooming SKU Detail")
    style_title(ws, 1, "Grooming SKU Detail", merge_to=8)
    style_subtitle(ws, 2, "All grooming service SKUs ranked by revenue", merge_to=8)

    row = 4
    headers = ["Rank", "SKU", "Service Name", "Service Type", "Units", "Net Sales", "% of Service Mix", "Avg Ticket"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 8)

    for rank, (_, r) in enumerate(sku_df.head(50).iterrows(), 1):
        row += 1
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=str(r["sku"]))
        ws.cell(row=row, column=3, value=r["name"])
        ws.cell(row=row, column=4, value=r["service_type"])
        ws.cell(row=row, column=5, value=fmt_int(r["units"]))
        ws.cell(row=row, column=6, value=fmt_currency(r["net_sales"]))
        ws.cell(row=row, column=7, value=fmt_pct(r["pct_mix"]))
        ws.cell(row=row, column=8, value=fmt_currency(r["avg_ticket"]))

    zebra_rows(ws, 5, row, 8)
    auto_width(ws)
    return ws


def write_top_retail(wb, top_rev, top_units):
    for sheet_name, data, sort_col in [
        ("Top 50 Retail (by $)", top_rev, "net_sales"),
        ("Top 50 Retail (by Units)", top_units, "units"),
    ]:
        ws = wb.create_sheet(sheet_name)
        style_title(ws, 1, sheet_name, merge_to=8)
        style_subtitle(ws, 2, f"Top 50 retail SKUs ranked by {'net sales' if '$' in sheet_name else 'units sold'}", merge_to=8)

        row = 4
        headers = ["Rank", "SKU", "Product Name", "Category", "Units", "Net Sales", "% of Retail", "Avg Price"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, 8)

        for rank, (_, r) in enumerate(data.iterrows(), 1):
            row += 1
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=str(r["sku"]))
            ws.cell(row=row, column=3, value=r["name"])
            ws.cell(row=row, column=4, value=r["retail_category"])
            ws.cell(row=row, column=5, value=fmt_int(r["units"]))
            ws.cell(row=row, column=6, value=fmt_currency(r["net_sales"]))
            ws.cell(row=row, column=7, value=fmt_pct(r["pct_retail"]))
            ws.cell(row=row, column=8, value=fmt_currency(r["avg_price"]))

        zebra_rows(ws, 5, row, 8)
        auto_width(ws)


def write_monthly_performance(wb, monthly_df, *, store_name, start_date, end_date):
    ws = wb.create_sheet("Monthly Performance")
    style_title(ws, 1, "Monthly Performance Dashboard", merge_to=9)
    style_subtitle(ws, 2, f"{store_name} -- {start_date} to {end_date}", merge_to=9)

    row = 4
    ws.cell(row=row, column=1, value="Total Store Revenue by Month").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    headers = ["Month", "Net Revenue", "Grooming Rev", "Retail Rev", "Groom %", "Retail %", "Tickets", "Avg Ticket", "MoM Growth"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 9)

    for _, r in monthly_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["month_name"])
        ws.cell(row=row, column=2, value=fmt_currency(r["net_revenue"]))
        ws.cell(row=row, column=3, value=fmt_currency(r["grooming_rev"]))
        ws.cell(row=row, column=4, value=fmt_currency(r["retail_rev"]))
        ws.cell(row=row, column=5, value=fmt_pct(r["groom_pct"]))
        ws.cell(row=row, column=6, value=fmt_pct(r["retail_pct"]))
        ws.cell(row=row, column=7, value=int(r["tickets"]))
        ws.cell(row=row, column=8, value=fmt_currency(r["avg_ticket"]))
        ws.cell(row=row, column=9, value=fmt_pct(r["mom_growth"]) if not pd.isna(r["mom_growth"]) else "--")

    zebra_rows(ws, 7, row, 9)

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=fmt_currency(monthly_df["net_revenue"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=3, value=fmt_currency(monthly_df["grooming_rev"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=4, value=fmt_currency(monthly_df["retail_rev"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=7, value=int(monthly_df["tickets"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=8, value=fmt_currency(monthly_df["net_revenue"].sum()/monthly_df["tickets"].sum())).font = Font(bold=True)

    if len(monthly_df) >= 2:
        jan = monthly_df.iloc[0]["net_revenue"]
        dec = monthly_df.iloc[-1]["net_revenue"]
        if jan > 0:
            growth = (dec - jan) / jan * 100
            row += 2
            ws.cell(row=row, column=1, value=f"Full year growth: {growth:.1f}% from {monthly_df.iloc[0]['month_name']} to {monthly_df.iloc[-1]['month_name']}. Avg ticket grew from {fmt_currency(monthly_df.iloc[0]['avg_ticket'])} to {fmt_currency(monthly_df.iloc[-1]['avg_ticket'])}.").font = Font(name="Inter", size=10, italic=True)

    auto_width(ws)

    try:
        chart = LineChart()
        chart.title = "Monthly Revenue Trend"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Month"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        data_start_row = row + 3
        for i, (_, r) in enumerate(monthly_df.iterrows()):
            ws.cell(row=data_start_row + i, column=11, value=r["month_name"])
            ws.cell(row=data_start_row + i, column=12, value=r["net_revenue"])
            ws.cell(row=data_start_row + i, column=13, value=r["grooming_rev"])
            ws.cell(row=data_start_row + i, column=14, value=r["retail_rev"])

        cats = Reference(ws, min_col=11, min_row=data_start_row, max_row=data_start_row + 11)
        data_ref = Reference(ws, min_col=12, max_col=14, min_row=data_start_row - 1, max_row=data_start_row + 11)

        ws.cell(row=data_start_row - 1, column=12, value="Total")
        ws.cell(row=data_start_row - 1, column=13, value="Grooming")
        ws.cell(row=data_start_row - 1, column=14, value="Retail")

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row + 3}")
    except Exception as e:
        print(f"  Chart error: {e}")

    return ws


def write_top_customers(wb, top_cust):
    ws = wb.create_sheet("Top Customers")
    style_title(ws, 1, "Top Customers -- Ranked by Total Spend", merge_to=7)
    style_subtitle(ws, 2, "Top 50 customers by total spend", merge_to=7)

    row = 4
    headers = ["Customer ID", "Total Spend", "Transactions", "Visit Days", "Spend/Visit", "Avg Txn", "Tier"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 7)

    tier_colors = {
        "Premium": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "High": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "Standard": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        "Value": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    }

    for _, r in top_cust.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=int(r["customer_id"]))
        ws.cell(row=row, column=2, value=fmt_currency(r["total_spend"]))
        ws.cell(row=row, column=3, value=int(r["transactions"]))
        ws.cell(row=row, column=4, value=int(r["visit_days"]))
        ws.cell(row=row, column=5, value=fmt_currency(r["spend_per_visit"]))
        ws.cell(row=row, column=6, value=fmt_currency(r["avg_txn"]))
        ws.cell(row=row, column=7, value=r["tier"])
        if r["tier"] in tier_colors:
            ws.cell(row=row, column=7).fill = tier_colors[r["tier"]]

    auto_width(ws)
    return ws


def write_customer_intelligence(wb, ci, *, store_name, start_date, end_date):
    ws = wb.create_sheet("Customer Intelligence")
    style_title(ws, 1, "Customer Intelligence Dashboard", merge_to=8)
    style_subtitle(ws, 2, f"{store_name} -- {start_date} to {end_date}", merge_to=8)

    row = 4
    ws.cell(row=row, column=1, value="Ticket Composition Analysis").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    for i, h in enumerate(["Ticket Type", "Tickets", "% of Tickets", "Revenue", "% of Revenue", "Avg Ticket"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 6)

    for _, r in ci["ticket_composition"].iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["ticket_type"])
        ws.cell(row=row, column=2, value=fmt_int(r["tickets"]))
        ws.cell(row=row, column=3, value=fmt_pct(r["pct_tickets"]))
        ws.cell(row=row, column=4, value=fmt_currency(r["revenue"]))
        ws.cell(row=row, column=5, value=fmt_pct(r["pct_revenue"]))
        ws.cell(row=row, column=6, value=fmt_currency(r["avg_ticket"]))

    row += 3
    ws.cell(row=row, column=1, value="Customer Visit Frequency").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    for i, h in enumerate(["Visit Frequency", "Customers", "% of Customers", "Revenue", "% of Revenue", "Avg Spend/Customer"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 6)

    freq_order = ["1 visit (trial)", "2 visits", "3-4 visits (developing)", "5-8 visits (loyal)", "9-12 visits (committed)", "13+ visits (champions)"]
    freq_df = ci["visit_frequency"]
    freq_df["_order"] = freq_df["bucket"].map({s: i for i, s in enumerate(freq_order)})
    freq_df = freq_df.sort_values("_order")

    for _, r in freq_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["bucket"])
        ws.cell(row=row, column=2, value=fmt_int(r["customers"]))
        ws.cell(row=row, column=3, value=fmt_pct(r["pct_customers"]))
        ws.cell(row=row, column=4, value=fmt_currency(r["revenue"]))
        ws.cell(row=row, column=5, value=fmt_pct(r["pct_revenue"]))
        ws.cell(row=row, column=6, value=fmt_currency(r["avg_spend"]))

    row += 3
    ws.cell(row=row, column=1, value="Customer Recency & Churn Risk").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    style_subtitle(ws, row + 1, "Days since last visit as of Dec 31, 2025", merge_to=5)
    row += 2
    for i, h in enumerate(["Segment", "Customers", "%", "Total Spend", "Avg Spend"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 5)

    seg_order = ["Active (last 30 days)", "Recent (31-60 days)", "At Risk (61-90 days)", "Lapsing (91-180 days)", "Lost (180+ days)"]
    churn_df = ci["churn_risk"]
    churn_df["_order"] = churn_df["segment"].map({s: i for i, s in enumerate(seg_order)})
    churn_df = churn_df.sort_values("_order")

    for _, r in churn_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["segment"])
        ws.cell(row=row, column=2, value=fmt_int(r["customers"]))
        ws.cell(row=row, column=3, value=fmt_pct(r["pct"]))
        ws.cell(row=row, column=4, value=fmt_currency(r["total_spend"]))
        ws.cell(row=row, column=5, value=fmt_currency(r["avg_spend"]))

    row += 3
    ws.cell(row=row, column=1, value="Day of Week Performance").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    for i, h in enumerate(["Day", "Revenue", "% of Week", "Tickets", "Avg Ticket", "Grooming Rev", "Retail Rev", "Retail %"], 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 8)

    for _, r in ci["day_of_week"].iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["day"])
        ws.cell(row=row, column=2, value=fmt_currency(r["revenue"]))
        ws.cell(row=row, column=3, value=fmt_pct(r["pct_week"]))
        ws.cell(row=row, column=4, value=int(r["tickets"]))
        ws.cell(row=row, column=5, value=fmt_currency(r["avg_ticket"]))
        ws.cell(row=row, column=6, value=fmt_currency(r["grooming_rev"]))
        ws.cell(row=row, column=7, value=fmt_currency(r["retail_rev"]))
        ws.cell(row=row, column=8, value=fmt_pct(r["retail_pct"]))

    auto_width(ws)
    return ws


def write_team_performance(wb, team_df):
    ws = wb.create_sheet("Team Performance")
    style_title(ws, 1, "Team Performance", merge_to=12)
    style_subtitle(ws, 2, "Groomer scorecards and productivity metrics", merge_to=12)

    row = 4
    ws.cell(row=row, column=1, value="Groomer Scorecard").font = Font(name="Inter", size=12, bold=True, color=TEDDY_BROWN)
    row += 2
    headers = ["Groomer", "Revenue", "Appointments", "$/Appt", "Core Visits",
               "SPA %", "Add-On %", "Teeth %", "Hours", "$/Hour"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(headers))

    for _, r in team_df.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=r["salesperson"])
        ws.cell(row=row, column=2, value=fmt_currency(r["revenue"]))
        ws.cell(row=row, column=3, value=int(r["appointments"]))
        ws.cell(row=row, column=4, value=fmt_currency(r["per_appt"]))
        ws.cell(row=row, column=5, value=fmt_int(r["core_visits"]))
        ws.cell(row=row, column=6, value=fmt_pct(r["spa_pct"]))
        ws.cell(row=row, column=7, value=fmt_pct(r["addon_pct"]))
        ws.cell(row=row, column=8, value=fmt_pct(r["teeth_pct"]))
        ws.cell(row=row, column=9, value=f"{r['hours']:.1f}" if not pd.isna(r.get("hours")) else "N/A")
        ws.cell(row=row, column=10, value=fmt_currency(r["per_hour"]) if not pd.isna(r.get("per_hour")) else "N/A")

        spa_val = r["spa_pct"]
        if spa_val >= 40:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif spa_val >= 25:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        else:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    zebra_rows(ws, 7, row, len(headers))

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=fmt_currency(team_df["revenue"].sum())).font = Font(bold=True)
    ws.cell(row=row, column=3, value=int(team_df["appointments"].sum())).font = Font(bold=True)

    auto_width(ws)
    return ws


def write_category_top10s(wb, cat_data):
    ws = wb.create_sheet("Category Top 10s")
    style_title(ws, 1, "Category Top 10s", merge_to=10)
    style_subtitle(ws, 2, "Top 10 by revenue (left) and by units (right) for each retail category", merge_to=10)

    row = 4
    for cat_name in sorted(cat_data.keys(), key=lambda x: cat_data[x]["total"], reverse=True):
        info = cat_data[cat_name]
        ws.cell(row=row, column=1, value=f"{cat_name}  |  Total Revenue: {fmt_currency(info['total'])}").font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        row += 1

        for i, h in enumerate(["Rank", "Product Name", "Revenue", "Units"], 1):
            ws.cell(row=row, column=i, value=h)
        for i, h in enumerate(["Rank", "Product Name", "Units", "Revenue"], 1):
            ws.cell(row=row, column=i + 5, value=h)
        style_header(ws, row, 4)
        for c in range(6, 10):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
            ws.cell(row=row, column=c).font = Font(name="Inter", size=10, bold=True, color=WHITE)

        top_rev = info["top_by_revenue"]
        top_units = info["top_by_units"]

        max_rows = max(len(top_rev), len(top_units))
        for i in range(max_rows):
            row += 1
            if i < len(top_rev):
                r = top_rev.iloc[i]
                ws.cell(row=row, column=1, value=i + 1)
                ws.cell(row=row, column=2, value=r["name"])
                ws.cell(row=row, column=3, value=fmt_currency(r["revenue"]))
                ws.cell(row=row, column=4, value=fmt_int(r["units"]))
            if i < len(top_units):
                r = top_units.iloc[i]
                ws.cell(row=row, column=6, value=i + 1)
                ws.cell(row=row, column=7, value=r["name"])
                ws.cell(row=row, column=8, value=fmt_int(r["units"]))
                ws.cell(row=row, column=9, value=fmt_currency(r["revenue"]))

        row += 3

    auto_width(ws)
    return ws


def write_brand_breakdown(wb, brand_data):
    ws = wb.create_sheet("Brand Breakdown")
    style_title(ws, 1, "Brand Breakdown by Category", merge_to=7)
    style_subtitle(ws, 2, "Top 10 brands per retail category, sorted by revenue", merge_to=7)

    row = 4
    for cat_name in sorted(brand_data.keys(), key=lambda x: brand_data[x]["total"], reverse=True):
        info = brand_data[cat_name]
        ws.cell(row=row, column=1, value=f"{cat_name}  |  Total: {fmt_currency(info['total'])}").font = Font(name="Inter", size=11, bold=True, color=TEDDY_BROWN)
        row += 1

        headers = ["Brand", "Net Sales", "% of Category", "Units", "SKU Count", "Avg Price/Item", "Revenue/SKU"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, 7)

        for _, r in info["brands"].iterrows():
            row += 1
            ws.cell(row=row, column=1, value=r["brand"])
            ws.cell(row=row, column=2, value=fmt_currency(r["net_sales"]))
            ws.cell(row=row, column=3, value=fmt_pct(r["pct_category"]))
            ws.cell(row=row, column=4, value=fmt_int(r["units"]))
            ws.cell(row=row, column=5, value=int(r["sku_count"]))
            ws.cell(row=row, column=6, value=fmt_currency(r["avg_price"]))
            ws.cell(row=row, column=7, value=fmt_currency(r["revenue_per_sku"]))

        row += 3

    auto_width(ws)
    return ws
